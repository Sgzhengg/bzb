"""
历史中标结果 API 接口

端点:
  GET    /api/v1/awards              中标结果列表（筛选/分页）
  GET    /api/v1/awards/{id}         单条详情
  DELETE /api/v1/awards/{id}         删除单条记录
  GET    /api/v1/awards/stats        中标统计概览
"""

import logging
import uuid
import asyncio
import os
import sys
import json
import re
from datetime import datetime, date
from typing import Optional, Dict

from fastapi import APIRouter, Depends, HTTPException, Query, BackgroundTasks
from sqlalchemy import select, func, and_, desc
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.session import get_db, AsyncSessionLocal
from app.models.historical_award import HistoricalAward
from app.models.client_relation import Purchaser

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/awards", tags=["中标结果"])

# ── 采集进度追踪（内存中，重启丢失） ──
_fetch_tasks: Dict[str, dict] = {}  # {task_id: {status, progress, message, ...}}


def _compute_eta(task: dict) -> dict:
    """根据已用时间和进度百分比，动态估算剩余时间"""
    import time as _time
    started = task.get("started_at")
    progress = task.get("progress", 0)
    if not started or progress <= 0 or progress >= 100:
        return {}
    try:
        elapsed = _time.time() - datetime.fromisoformat(started).timestamp()
        if elapsed < 2:
            return {}
        eta_total = elapsed / (progress / 100.0)
        remaining = max(0, eta_total - elapsed)
        return {"elapsed_seconds": round(elapsed), "eta_seconds": round(remaining)}
    except Exception:
        return {}


async def _classify_award(title: str, content: str = "", data_source: str = "") -> dict:
    """复用公告采集的核心二步管线：keyword_filter → classify_and_extract。
    跳过 _is_mobile_purchaser 和 中标公示过滤（中标结果本就含这些词）。
    """
    import sys as _sys, os as _os
    _backend = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
    _sys.path.insert(0, _backend)

    content = content or ""

    # ── Step 1: keyword_filter（仅用安全词+排除词，不检查采购单位）──
    def _step1():
        try:
            from app.services.keyword_filter import (
                SAFETY_KEYWORDS, HARD_EXCLUDE_KEYWORDS, _match_keywords
            )
        except ImportError:
            from services.keyword_filter import (
                SAFETY_KEYWORDS, HARD_EXCLUDE_KEYWORDS, _match_keywords
            )
        combined = f"{title} {content}"
        excluded = _match_keywords(combined, HARD_EXCLUDE_KEYWORDS)
        if excluded:
            return False, "", f"命中排除词: {excluded[0]}"
        matched = _match_keywords(combined, SAFETY_KEYWORDS)
        if not matched:
            return False, "", "未命中广告安全词"
        return True, matched, ""

    is_ad, matched, reason = await asyncio.to_thread(_step1)
    if not is_ad:
        return {"is_ad": False, "category": "", "reason": reason}

    # ── Step 2: LLM 精确分类 ──
    try:
        def _step2():
            try:
                from app.services.llm_classifier import classify_and_extract
            except ImportError:
                from services.llm_classifier import classify_and_extract
            return classify_and_extract(title, content)

        unified = await asyncio.to_thread(_step2)
        llm_category = unified.get("category", "")
        llm_budget = unified.get("budget")
        is_ad = True
        category = llm_category or ""
        budget = llm_budget
    except Exception as e:
        logger.warning(f"LLM step failed, using keyword result: {e}")
        is_ad, category, budget = True, "", None

    # ── Step 3: 非广告项目 → industry_classifier 兜底 ──
    if not is_ad or not category:
        try:
            def _step3():
                try:
                    from app.services.industry_classifier import classify_industry_and_category
                except ImportError:
                    from services.industry_classifier import classify_industry_and_category
                return classify_industry_and_category(title, content, data_source)
            ind_result = await asyncio.to_thread(_step3)
            cat = ind_result.get("project_category", "")
            if cat and cat != "其他采购":
                category = cat
                is_ad = True  # 有具体类别就入库
                logger.info(f"  🔄 industry_classifier: [{cat}] {title[:50]}")
        except Exception as e:
            logger.warning(f"industry_classifier fallback failed: {e}")

    # ── Step 4: LLM 折扣率提取（从PDF正文中提取中标折扣率/金额）──
    discount_rate = None
    if content and len(content) > 100:
        try:
            def _extract_discount():
                import httpx as _hx
                from app.core.config import settings as _cfg
                if not _cfg.LLM_API_KEY:
                    return None
                prompt = f"""从以下中标公示正文中提取中标折扣率或中标金额。
折扣率格式如"折扣率85%"、"中标折扣82.5%"、"应答折扣率90%"等。
金额格式如"中标金额123.5万元"。
仅回复JSON：{{"discount_rate": 数字(百分比,如85.5), "amount": 数字(万元)}}，未找到则null。

正文：{content[:3000]}"""
                with _hx.Client(timeout=15) as client:
                    resp = client.post(
                        f"{_cfg.LLM_API_BASE}/chat/completions",
                        headers={"Authorization": f"Bearer {_cfg.LLM_API_KEY}", "Content-Type": "application/json"},
                        json={"model": _cfg.LLM_MODEL, "temperature": 0, "max_tokens": 100,
                              "messages": [{"role": "user", "content": prompt}]},
                    )
                    text = resp.json()["choices"][0]["message"]["content"]
                    import re as _re, json as _json
                    m = _re.search(r'\{[^}]+\}', text)
                    if m:
                        d = _json.loads(m.group())
                        return d.get("discount_rate") or d.get("amount")
                return None
            discount_rate = await asyncio.to_thread(_extract_discount)
        except Exception:
            pass

    return {
        "is_ad": is_ad,
        "category": category,
        "budget": budget,
        "discount_rate": discount_rate,
        "reason": "",
    }


async def _import_to_db(backend_dir: str, adapter: str, province: str = "") -> int:
    """将 crawl_winning_results.py 输出的 JSON 导入 historical_awards 表。

    Returns:
        成功导入的记录数。
    """
    output_dir = os.path.join(backend_dir, "output")
    province_slug = province.replace(",", "_") if province else "quanguo"
    json_path = os.path.join(output_dir, f"winning_results_{adapter}_{province_slug}.json")

    if not os.path.exists(json_path):
        logger.warning(f"⚠️ 采集结果文件不存在: {json_path}")
        return 0

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # 提取所有 items
    all_items = []
    adapters_data = data.get("adapters", [])
    for adp_data in adapters_data:
        all_items.extend(adp_data.get("items", []))

    if not all_items:
        logger.info("📭 没有可导入的中标结果")
        return 0

    imported = 0
    async with AsyncSessionLocal() as session:
        # 确保默认采购方存在
        from sqlalchemy import select as sa_select
        existing = await session.execute(
            sa_select(Purchaser).where(Purchaser.id == 1)
        )
        if not existing.scalar_one_or_none():
            session.add(Purchaser(
                id=1,
                name="中国移动通信集团广东有限公司",
                level="省公司",
                region="广州",
            ))
            await session.flush()

        for item in all_items:
            title = (item.get("title") or "").strip()
            if not title:
                continue

            source_url = (item.get("source_url") or "").strip()

            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 修复：提前提取 winner_name，供后续去重检查使用
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            winner_name = item.get("winner_name", "").strip()
            if not winner_name:
                winner_name = _extract_winner_from_title(title)

            # 检查重复：同一公告同一中标方只入库一次
            if source_url and winner_name:
                dup_check = await session.execute(
                    sa_select(HistoricalAward).where(
                        HistoricalAward.source_url == source_url,
                        HistoricalAward.winner_name == winner_name,
                    )
                )
                if dup_check.scalar_one_or_none():
                    continue
            elif source_url:
                dup_check = await session.execute(
                    sa_select(HistoricalAward).where(
                        HistoricalAward.source_url == source_url
                    )
                )
                if dup_check.scalar_one_or_none():
                    continue
            elif winner_name:
                dup_check = await session.execute(
                    sa_select(HistoricalAward).where(
                        HistoricalAward.project_name == title[:500],
                        HistoricalAward.winner_name == winner_name,
                    )
                )
                if dup_check.scalar_one_or_none():
                    continue

            # 解析日期
            pub_date_str = item.get("publish_date", "")
            bid_open_date = date.today()
            if pub_date_str:
                for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
                    try:
                        bid_open_date = datetime.strptime(pub_date_str[:10], fmt).date()
                        break
                    except ValueError:
                        continue

            # 解析折扣率（先取爬虫值，后面LLM可能覆盖）
            discount_rate = item.get("discount_rate")
            if discount_rate is not None:
                try:
                    discount_rate = float(discount_rate)
                except (ValueError, TypeError):
                    discount_rate = None
            else:
                discount_rate = None

            # 确定数据来源
            data_source = item.get("data_source", "") or item.get("adapter", "") or adapter

            # 确定项目类别（先用关键词猜测，LLM 有结果则覆盖）
            project_category = _guess_category(title)

            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 第二步：关键词+LLM 二步法鉴定（对标公告采集管线）
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            try:
                classify_result = await _classify_award(title, item.get("original_content", ""), data_source)
                if not classify_result.get("is_ad"):
                    # 非广告项目也有类别（industry_classifier 已处理），继续入库
                    llm_category = classify_result.get("category", "")
                    if llm_category:
                        project_category = llm_category
                        logger.info(f"  ✅ 非广告类 [{project_category}]: {title[:60]}")
                    else:
                        logger.info(f"  ⏭️ 无法分类，跳过: {title[:60]}")
                        continue
                # LLM category 覆盖关键词猜测（但仅当LLM类别为广告类赛道时）
                llm_category = classify_result.get("category", "")
                AD_CATEGORIES = {"广告类", "广告创意设计", "物料制作印刷", "活动策划执行",
                    "品牌宣传传播", "视频内容制作", "新媒体运营", "新媒体运营类", "媒介资源投放", "媒介投放类",
                    "渠道营销推广", "渠道营销类", "内容制作类", "创意设计类", "其他",
                    "通信工程建设", "ICT系统集成", "设备采购", "网络维护代维", "行政物业", "设计勘察"}
                if llm_category in AD_CATEGORIES:
                    project_category = llm_category
                # LLM discount_rate
                llm_discount = classify_result.get("discount_rate")
                if not discount_rate and llm_discount:
                    try:
                        discount_rate = float(llm_discount)
                    except (ValueError, TypeError):
                        pass
                # LLM budget
                llm_budget = classify_result.get("budget")
                if not discount_rate and llm_budget:
                    try:
                        discount_rate = float(llm_budget)
                    except (ValueError, TypeError):
                        pass
                logger.info(f"  ✅ 广告类 [{project_category}] 份额={discount_rate}: {title[:60]}")
            except Exception as e:
                logger.warning(f"  ⚠️ 分类异常，按关键词结果入库: {e}")

            try:
                award = HistoricalAward(
                    project_name=title[:500],
                    purchaser_id=1,
                    winner_name=winner_name or "未知中标方",
                    winner_type="头部常客",
                    bid_amount=0,
                    budget_amount=0,
                    discount_rate=discount_rate,
                    project_category=project_category or "其他",
                    bid_open_date=bid_open_date,
                    is_continuous=False,
                    continuous_count=0,
                    source_url=source_url,
                    data_source=data_source,
                )
                session.add(award)
                imported += 1
            except Exception as e:
                logger.warning(f"  ⚠️ 导入失败 [{title[:60]}]: {e}")

        await session.commit()

    logger.info(f"✅ 已导入 {imported} 条中标结果到 historical_awards")
    return imported


def _extract_winner_from_title(title: str) -> str:
    """从标题中提取中标方名称。"""
    patterns = [
        r'[（(]([^）)]+)[）)]\s*中标',
        r'中标(?:候选)?人?[：:]\s*([^\s，,。.]{2,30})',
        r'([^\s，,。.]{2,30}?(?:有限公司|有限责任公司|公司|集团))\s*(?:中标|中选|成交)',
        r'(?:拟|确定)\s*([^\s，,。.]{2,30}?(?:有限公司|有限责任公司|公司))\s*为',
    ]
    for pat in patterns:
        m = re.search(pat, title)
        if m:
            return m.group(1).strip()
    # 无中标方时尝试从标题提取公司名
    company_match = re.search(r'([^\s，,。.]{2,20}(?:有限公司|有限责任公司|公司|集团))', title)
    if company_match:
        return company_match.group(1).strip()
    return ""


def _guess_category(title: str) -> str:
    """根据标题推测项目类别（与 announcements 表对齐）。"""
    category_map = {
        "广告": "广告创意设计", "创意设计": "广告创意设计", "全案策划": "广告创意设计",
        "物料": "物料制作印刷", "喷绘": "物料制作印刷", "印刷": "物料制作印刷",
        "活动策划": "活动策划执行", "活动执行": "活动策划执行", "发布会": "活动策划执行", "路演": "活动策划执行", "展会": "活动策划执行",
        "品牌宣传": "品牌宣传传播", "品牌推广": "品牌宣传传播", "品牌传播": "品牌宣传传播", "宣传": "品牌宣传传播",
        "视频制作": "视频内容制作", "宣传片": "视频内容制作", "短视频": "视频内容制作",
        "新媒体": "新媒体运营", "公众号": "新媒体运营", "抖音": "新媒体运营", "直播": "新媒体运营",
        "媒介": "媒介资源投放", "投放": "媒介资源投放", "户外广告": "媒介资源投放", "电梯广告": "媒介资源投放",
        "渠道营销": "渠道营销推广", "地推": "渠道营销推广", "网格": "渠道营销推广", "触点": "渠道营销推广",
        "基站": "通信工程建设", "机房": "通信工程建设", "光缆": "通信工程建设", "铁塔": "通信工程建设", "管线": "通信工程建设", "土建": "通信工程建设", "通信工程": "通信工程建设", "施工": "通信工程建设",
        "系统集成": "ICT系统集成", "ICT": "ICT系统集成", "软件开发": "ICT系统集成", "平台": "ICT系统集成", "大数据": "ICT系统集成", "云计算": "ICT系统集成", "软硬件": "ICT系统集成",
        "服务器": "设备采购", "交换机": "设备采购", "路由器": "设备采购", "设备采购": "设备采购", "终端": "设备采购",
        "网络维护": "网络维护代维", "代维": "网络维护代维", "运维": "网络维护代维", "维保": "网络维护代维", "网络优化": "网络维护代维", "网络安全": "网络维护代维",
        "物业": "行政物业", "保安": "行政物业", "保洁": "行政物业", "食堂": "行政物业", "消防": "行政物业", "快递": "行政物业", "速递": "行政物业",
        "勘察": "设计勘察", "可行性研究": "设计勘察", "规划": "设计勘察",
        "制作": "视频内容制作", "设计": "广告创意设计",
    }
    for kw, cat in category_map.items():
        if kw in title:
            return cat
    return "其他采购"


@router.get("", summary="获取中标结果列表")
async def list_awards(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    project_category: Optional[str] = Query(None, description="项目类别筛选"),
    winner_type: Optional[str] = Query(None, description="中标方类型"),
    purchaser_id: Optional[int] = Query(None, description="采购方ID"),
    search: Optional[str] = Query(None, description="项目名称/中标方搜索"),
    data_source: Optional[str] = Query(None, description="数据来源: b2b_10086(移动)/telecom(电信)/unicom(联通)/gd_zbtb/gd_ygp"),
    collected_from: Optional[str] = Query(None, description="采集时间起 (YYYY-MM-DD)"),
    collected_to: Optional[str] = Query(None, description="采集时间止 (YYYY-MM-DD)"),
    db: AsyncSession = Depends(get_db),
):
    """获取历史中标结果列表，支持筛选和分页。"""
    conditions = []

    if project_category:
        conditions.append(HistoricalAward.project_category == project_category)
    if collected_from:
        conditions.append(HistoricalAward.created_at >= collected_from)
    if collected_to:
        conditions.append(HistoricalAward.created_at < collected_to + "T23:59:59")
    if winner_type:
        conditions.append(HistoricalAward.winner_type == winner_type)
    if purchaser_id:
        conditions.append(HistoricalAward.purchaser_id == purchaser_id)
    if data_source:
        conditions.append(HistoricalAward.data_source == data_source)
    if search:
        conditions.append(
            HistoricalAward.project_name.ilike(f"%{search}%")
            | HistoricalAward.winner_name.ilike(f"%{search}%")
        )

    # 总数
    count_q = select(func.count()).select_from(HistoricalAward)
    if conditions:
        count_q = count_q.where(and_(*conditions))
    total = (await db.execute(count_q)).scalar() or 0

    # 列表
    list_q = (
        select(HistoricalAward)
        .order_by(desc(HistoricalAward.bid_open_date))
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    if conditions:
        list_q = list_q.where(and_(*conditions))

    result = await db.execute(list_q)
    awards = result.scalars().all()

    items = [
        {
            "id": a.id,
            "project_name": a.project_name,
            "purchaser_id": a.purchaser_id,
            "purchaser_name": a.purchaser.name if a.purchaser else "",
            "winner_name": a.winner_name,
            "winner_type": a.winner_type,
            "bid_amount": float(a.bid_amount) if a.bid_amount else None,
            "budget_amount": float(a.budget_amount) if a.budget_amount else None,
            "discount_rate": float(a.discount_rate) if a.discount_rate else None,
            "project_category": a.project_category,
            "bid_open_date": a.bid_open_date.isoformat() if a.bid_open_date else None,
            "contract_start": a.contract_start.isoformat() if a.contract_start else None,
            "contract_end": a.contract_end.isoformat() if a.contract_end else None,
            "is_continuous": a.is_continuous,
            "continuous_count": a.continuous_count,
            "source_url": a.source_url or "",
            "data_source": a.data_source or "",
        }
        for a in awards
    ]

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": items,
    }


@router.get("/stats", summary="中标统计概览")
async def award_stats(
    db: AsyncSession = Depends(get_db),
):
    """获取中标结果统计数据。"""
    # 总数
    total_q = select(func.count()).select_from(HistoricalAward)
    total = (await db.execute(total_q)).scalar() or 0

    # 总金额
    amount_q = select(func.sum(HistoricalAward.bid_amount)).select_from(HistoricalAward)
    total_amount = (await db.execute(amount_q)).scalar() or 0

    # 去重中标方数
    from sqlalchemy import distinct
    winner_count_q = select(func.count(distinct(HistoricalAward.winner_name))).select_from(HistoricalAward)
    winner_count = (await db.execute(winner_count_q)).scalar() or 0

    # 按项目类别统计
    cat_q = (
        select(
            HistoricalAward.project_category,
            func.count().label("count"),
        )
        .group_by(HistoricalAward.project_category)
        .order_by(desc("count"))
    )
    cat_result = await db.execute(cat_q)
    cat_stats = [
        {"category": row.project_category, "count": row.count}
        for row in cat_result.fetchall()
    ]

    return {
        "total": total,
        "total_amount": round(float(total_amount), 1),
        "winner_count": winner_count,
        "categories": cat_stats,
    }


@router.get("/export", summary="导出中标结果为Excel")
async def export_awards(
    db: AsyncSession = Depends(get_db),
    data_source: str = Query(None, description="数据源筛选: b2b_10086/telecom/unicom"),
):
    """将所有中标结果导出为 Excel 文件。"""
    from io import BytesIO
    from datetime import date
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    from urllib.parse import quote

    # 仅导出广告类中标结果
    AD_KEYWORDS = [
        "广告", "宣传", "品牌", "活动策划", "新媒体", "视频制作",
        "营销", "设计", "物料", "推广", "传播", "策划", "会展",
        "发布", "创意", "公关", "媒介", "视觉", "拍摄", "制作",
        "印刷", "展示", "展览", "路演", "地推", "促销",
    ]

    query = select(HistoricalAward).order_by(desc(HistoricalAward.bid_open_date))
    if data_source:
        query = query.where(HistoricalAward.data_source == data_source)
    result = await db.execute(query)
    all_awards = result.scalars().all()

    # 过滤非广告类
    awards = []
    for a in all_awards:
        title = a.project_name or ""
        if any(kw in title for kw in AD_KEYWORDS):
            awards.append(a)

    wb = Workbook()
    ws = wb.active
    ws.title = "中标结果"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center")
    data_align_center = Alignment(horizontal="center", vertical="center")
    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    headers = ["序号", "项目名称", "中标方", "中标份额(%)", "项目类别", "公示日期", "来源", "公示链接"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.freeze_panes = "A2"

    SOURCE_LABELS = {"b2b_10086": "中国移动", "telecom": "中国电信", "unicom": "中国联通"}

    for row_idx, award in enumerate(awards, 2):
        source_url = award.source_url or ""
        source_label = SOURCE_LABELS.get(award.data_source, award.data_source or "")

        row_data = [
            row_idx - 1,
            award.project_name or "",
            award.winner_name or "",
            float(award.discount_rate) if award.discount_rate else "",
            award.project_category or "其他",
            award.bid_open_date.strftime("%Y-%m-%d") if award.bid_open_date else "",
            source_label,
            source_url,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in (1, 4, 6, 7):
                cell.alignment = data_align_center
            elif col_idx == 8 and value:
                cell.value = "打开链接"
                cell.hyperlink = value
                cell.font = link_font
                cell.alignment = data_align_center
            elif col_idx == 2:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = even_fill

    col_widths = [6, 52, 18, 13, 13, 13, 10, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(awards) + 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"中标结果_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@router.get("/{award_id}", summary="获取中标结果详情")
async def get_award_detail(
    award_id: int,
    db: AsyncSession = Depends(get_db),
):
    """获取单条中标结果的详细信息。"""
    award = await db.get(HistoricalAward, award_id)
    if not award:
        raise HTTPException(status_code=404, detail=f"中标记录 {award_id} 不存在")

    return award.to_dict()


@router.post("/fetch", summary="手动触发中标结果采集")
async def fetch_awards(
    background_tasks: BackgroundTasks,
    province: Optional[str] = Query(None, description="目标省份，空为全国"),
    adapter: Optional[str] = Query(None, description="运营商: b2b_10086(移动)/telecom(电信)/unicom(联通)/all(全部)"),
    skip: Optional[bool] = Query(False, description="跳过爬虫，直接导入现有数据"),
):
    """后台触发中标结果数据采集（b2b.10086.cn），返回 task_id 用于轮询进度。"""
    task_id = str(uuid.uuid4())[:8]
    province_name = province or "全国"
    adapter_name = adapter or "b2b_10086"

    adapter_label_map = {
        "all": "全部运营商（移动+电信+联通）",
        "b2b_10086": "中国移动",
        "telecom": "中国电信",
        "unicom": "中国联通",
    }
    adapter_label = adapter_label_map.get(adapter_name, adapter_name)

    _fetch_tasks[task_id] = {
        "status": "starting",
        "progress": 0,
        "message": f"正在启动中标结果采集引擎（{adapter_label} × {province_name}）...",
        "province": province_name,
        "adapter": adapter_name,
        "started_at": datetime.now().isoformat(),
        "result_count": 0,
        "error": None,
    }

    async def _run():
        logger.info(f"🕷️ 中标结果采集开始 (adapter={adapter_name}, province={province_name}, task={task_id})...")
        try:
            import subprocess
            cwd = os.path.dirname(os.path.abspath(__file__))
            cwd = os.path.dirname(os.path.dirname(os.path.dirname(cwd)))  # -> backend/
            env = os.environ.copy()
            if sys.platform == "win32":
                env["PYTHONASYNCIODEFAULTLOOPPOLICY"] = "WindowsProactorEventLoopPolicy"

            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            # 检查是否有最近采集的 JSON 文件，如有则直接导入
            # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            province_slug = province.replace(",", "_") if province else "quanguo"
            json_path = os.path.join(cwd, "output", f"winning_results_{adapter_name}_{province_slug}.json")
            logger.info(f"[SKIP] 检查文件: {json_path}")

            skip_crawl = False
            file_age = 999999  # 默认值，表示文件不存在
            if os.path.exists(json_path):
                import time
                file_age = time.time() - os.path.getmtime(json_path)
                logger.info(f"[SKIP] 文件存在，年龄: {int(file_age/60)} 分钟")
                # 如果文件在 2 小时内生成，跳过爬虫直接导入
                if file_age < 7200:
                    skip_crawl = True
                    logger.info(f"[SKIP] 发现最近采集文件 ({int(file_age/60)} 分钟前)，跳过爬虫直接导入")
                else:
                    logger.info(f"[SKIP] 文件过旧 ({int(file_age/60)} 分钟)，需要重新采集")
            else:
                logger.info(f"[SKIP] 文件不存在，需要爬虫采集")

            if skip_crawl:
                _fetch_tasks[task_id].update(
                    status="running", progress=50,
                    message=f"发现最近数据，正在导入...",
                )
            else:
                _fetch_tasks[task_id].update(
                    status="running", progress=10,
                    message=f"正在搜索 {adapter_label} × {province_name} 中标公告...",
                )

                cmd = [sys.executable, "crawl_winning_results.py",
                       "--adapter", adapter_name]
                if province and province.strip():
                    cmd.extend(["--province", province])

                _fetch_tasks[task_id].update(
                    progress=30,
                    message=f"正在爬取 {adapter_label} × {province_name} 中标数据...",
                )

                proc = await asyncio.to_thread(
                    subprocess.run,
                    cmd,
                    capture_output=True, text=True, cwd=cwd, timeout=1200, env=env,
                )

                if proc.returncode != 0:
                    err_msg = proc.stderr[:500] if proc.stderr else f"退出码: {proc.returncode}"
                    _fetch_tasks[task_id].update(
                        status="failed", progress=0,
                        message=f"采集失败",
                        error=err_msg,
                    )
                    logger.error(f"❌ 中标结果采集失败 (task={task_id}): {err_msg}")
                    return

            _fetch_tasks[task_id].update(
                progress=70,
                message="正在将采集结果导入数据库...",
            )

            imported_count = await _import_to_db(cwd, adapter_name, province)

            _fetch_tasks[task_id].update(
                status="completed", progress=100,
                message=f"{adapter_label} × {province_name} 中标结果采集完成！共导入 {imported_count} 条",
                result_count=imported_count,
            )
            logger.info(f"✅ 中标结果采集完成 (task={task_id}): 导入 {imported_count} 条")

        except subprocess.TimeoutExpired as e:
            logger.error(f"❌ 中标结果采集超时 (task={task_id})")
            _fetch_tasks[task_id].update(
                status="failed", progress=0,
                message="采集超时（数据量较大，请稍后重试或使用现有数据）",
                error=f"超时: {str(e)}",
            )
        except Exception as e:
            logger.error(f"❌ 中标结果采集失败 (task={task_id}): {e}")
            _fetch_tasks[task_id].update(
                status="failed", progress=0,
                message=f"采集失败: {str(e)}",
                error=str(e),
            )

    background_tasks.add_task(_run)

    return {
        "status": "started",
        "task_id": task_id,
        "message": f"中标结果采集已在后台启动（{adapter_label} × {province_name}）",
    }


@router.get("/fetch/status/{task_id}", summary="查询中标采集进度")
async def get_fetch_status(task_id: str):
    """查询采集任务的实时进度。"""
    task = _fetch_tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"任务 {task_id} 不存在或已过期")
    eta = _compute_eta(task)
    return {**task, **eta}


@router.delete("/{award_id}", summary="删除中标结果")
async def delete_award(
    award_id: int,
    db: AsyncSession = Depends(get_db),
):
    """删除单条中标结果记录。"""
    award = await db.get(HistoricalAward, award_id)
    if not award:
        raise HTTPException(status_code=404, detail=f"中标记录 {award_id} 不存在")

    await db.delete(award)
    await db.commit()
    logger.info(f"已删除中标记录: id={award_id}, project={award.project_name}, winner={award.winner_name}")

    return {"ok": True, "message": f"已删除中标记录 {award_id}"}
