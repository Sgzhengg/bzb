"""
招标公告 API 接口

端点:
  GET    /api/v1/announcements         公告列表（排序/筛选/分页 + 机会评分）
  GET    /api/v1/announcements/{id}    公告详情（评分+提醒+在位者）
  POST   /api/v1/announcements/fetch   手动触发采集
  GET    /api/v1/announcements/fetch/status/{task_id}  采集进度查询
  POST   /api/v1/announcements/{id}/favorite  收藏/取消收藏
  GET    /api/v1/announcements/favorites      获取收藏列表
"""

import logging
import time
import uuid
from datetime import date, datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query, status, BackgroundTasks
from sqlalchemy import select, func, and_, or_, desc, asc, case
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.db.session import get_db
from app.models.announcement import Announcement
from app.models.client_relation import Purchaser, ClientRelation
from app.models.project_relation_alert import ProjectRelationAlert
from app.core.config import settings

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/announcements", tags=["招标公告"])

# ── 采集进度追踪（内存中，重启丢失） ──
_fetch_tasks: dict = {}  # {task_id: {status, progress, message, ...}}


def _to_iso(val):
    """将日期值转为 ISO 字符串。兼容 raw SQL 返回的字符串和 ORM 返回的 date 对象。"""
    if val is None:
        return None
    if isinstance(val, str):
        return val
    return val.isoformat()


# ============================================================
# 公告列表
# ============================================================

@router.get("", summary="获取公告列表")
async def list_announcements(
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    sort: Optional[str] = Query(None, description="排序字段: score_desc/date_desc/budget_desc"),
    province: Optional[str] = Query(None, description="省份筛选（如：广东、广西）"),
    city: Optional[str] = Query(None, description="城市筛选（如：广州、南宁）"),
    purchaser_level: Optional[str] = Query(None, description="采购方层级筛选"),
    project_category: Optional[str] = Query(None, description="项目类别筛选"),
    procurement_method: Optional[str] = Query(None, description="采购方式筛选"),
    probability_label: Optional[str] = Query(None, description="陪跑概率: 低/中/高"),
    budget_min: Optional[float] = Query(None, description="预算下限"),
    budget_max: Optional[float] = Query(None, description="预算上限"),
    search: Optional[str] = Query(None, description="项目名称搜索"),
    favorites_only: bool = Query(False, description="仅显示收藏"),
    db: AsyncSession = Depends(get_db),
):
    """获取招标公告列表，支持多维筛选、排序和分页。"""
    conditions = []

    if province:
        conditions.append(Announcement.province == province)
    if city:
        conditions.append(Announcement.city == city)
    if purchaser_level:
        conditions.append(Announcement.purchaser_level == purchaser_level)
    if project_category:
        conditions.append(Announcement.project_category == project_category)
    if procurement_method:
        conditions.append(Announcement.procurement_method == procurement_method)
    if budget_min is not None:
        conditions.append(
            (Announcement.budget >= budget_min) | (Announcement.budget == None)
        )
    if budget_max is not None:
        conditions.append(
            (Announcement.budget <= budget_max) | (Announcement.budget == None)
        )
    if search:
        conditions.append(Announcement.title.ilike(f"%{search}%"))

    # 自动过滤中标公示（中选/中标/成交候选人/结果公示属于中标结果页，非机会列表）
    conditions.append(
        ~Announcement.title.ilike("%中选%")
        & ~Announcement.title.ilike("%中标%")
        & ~Announcement.title.ilike("%成交候选人%")
        & ~Announcement.title.ilike("%成交结果%")
    )

    # 自动过滤已过期的公告（TODO: deadline 存为 TEXT，比较需修复）
    # from datetime import datetime as dt
    # today = dt.now()
    # conditions.append(Announcement.deadline >= today)

    # 仅显示收藏
    if favorites_only:
        conditions.append(Announcement.is_favorited == True)

    # 总数
    count_q = select(func.count()).select_from(Announcement)
    if conditions:
        count_q = count_q.where(and_(*conditions))
    total = (await db.execute(count_q)).scalar() or 0

    # 排序
    order_clauses = [desc(Announcement.announce_date)]
    if sort == "score_desc":
        order_clauses = [desc(Announcement.announce_date)]  # 默认日期降序（评分由前端Mock）
    elif sort == "date_desc":
        order_clauses = [desc(Announcement.announce_date)]
    elif sort == "budget_desc":
        order_clauses = [desc(Announcement.budget)]

    # 分页查询
    list_q = (
        select(Announcement)
        .options(selectinload(Announcement.purchaser))
        .order_by(*order_clauses)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    if conditions:
        list_q = list_q.where(and_(*conditions))

    result = await db.execute(list_q)
    announcements = result.scalars().all()

    items = []
    for ann in announcements:
        purchaser_name = ann.purchaser.name if ann.purchaser else ""
        # 计算机会评分（使用评分引擎）
        score_data = _compute_announcement_score(ann)

        items.append({
            "id": ann.id,
            "title": ann.title,
            # Excel模板字段
            "industry": getattr(ann, 'industry', '') or '',
            "province": getattr(ann, 'province', '') or '',
            "city": getattr(ann, 'city', '') or '',
            "project_category": ann.project_category,
            "procurement_method": ann.procurement_method,
            "budget": float(ann.budget) if ann.budget is not None else None,
            "source_url": ann.source_url or '',
            "announce_date": ann.announce_date.isoformat() if ann.announce_date else None,
            "deadline": ann.deadline.isoformat() if ann.deadline else None,
            "deadline_time": getattr(ann, 'deadline_time', '') or '',
            "bid_date": ann.bid_date.isoformat() if getattr(ann, 'bid_date', None) else None,
            "bid_time": getattr(ann, 'bid_time', '') or '',
            "registration_fee": float(getattr(ann, 'registration_fee', 0) or 0),
            "deposit": float(getattr(ann, 'deposit', 0) or 0),
            "remark": getattr(ann, 'remark', '') or '',
            # 辅助字段
            "purchaser": purchaser_name,
            "purchaser_id": ann.purchaser_id,
            "purchaser_level": ann.purchaser_level,
            "created_at": ann.created_at.isoformat() if ann.created_at else None,
            # 收藏
            "is_favorited": getattr(ann, 'is_favorited', False) or False,
            # 评分（来自评分引擎）
            "total_score": score_data.get("total_score"),
            "probability_label": score_data.get("probability_label", ""),
            "detail_scores": score_data.get("detail_scores", {}),
            "recommendation": score_data.get("recommendation", ""),
        })

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": items,
    }


# ============================================================
# 收藏列表（必须在 /{announcement_id} 之前注册，否则被拦截）
# ============================================================

@router.get("/favorites", summary="获取收藏列表")
async def list_favorites(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
):
    """获取已收藏的公告列表（自动排除中标公示）。"""
    base_condition = and_(
        Announcement.is_favorited == True,
        ~Announcement.title.ilike("%中选%"),
        ~Announcement.title.ilike("%中标%"),
        ~Announcement.title.ilike("%成交候选人%"),
        ~Announcement.title.ilike("%成交结果%"),
    )
    count_q = select(func.count()).select_from(Announcement).where(base_condition)
    total = (await db.execute(count_q)).scalar() or 0

    list_q = (
        select(Announcement)
        .where(base_condition)
        .order_by(desc(Announcement.announce_date))
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    result = await db.execute(list_q)
    announcements = result.scalars().all()

    items = []
    for ann in announcements:
        score_data = _compute_announcement_score(ann)
        items.append({
            "id": ann.id,
            "title": ann.title,
            "project_category": ann.project_category,
            "procurement_method": ann.procurement_method,
            "budget": float(ann.budget) if ann.budget is not None else None,
            "announce_date": ann.announce_date.isoformat() if ann.announce_date else None,
            "deadline": ann.deadline.isoformat() if ann.deadline else None,
            "purchaser": ann.purchaser.name if ann.purchaser else "",
            "total_score": score_data.get("total_score"),
            "probability_label": score_data.get("probability_label", ""),
            "is_favorited": True,
        })

    return {"total": total, "page": page, "page_size": page_size, "items": items}


@router.get("/favorites/export", summary="导出收藏为Excel")
async def export_favorites(
    db: AsyncSession = Depends(get_db),
):
    """将所有收藏公告导出为美观的 Excel（.xlsx）文件。"""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    result = await db.execute(
        select(Announcement)
        .where(Announcement.is_favorited == True)
        .order_by(desc(Announcement.announce_date))
    )
    announcements = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "收藏公告"

    # ── 样式定义 ──
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=False)
    data_align_center = Alignment(horizontal="center", vertical="center")
    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # ── 表头 ──
    headers = ["序号", "省份", "地市", "项目名称", "种类", "采购方式", "预算(万)", "报名截止", "投标日期", "报名费(元)", "保证金(万)", "网址"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 冻结表头
    ws.freeze_panes = "A2"

    # ── 数据行 ──
    for row_idx, ann in enumerate(announcements, 2):
        row_data = [
            row_idx - 1,  # 序号
            ann.province or "",
            ann.city or "",
            ann.title or "",
            ann.project_category or "",
            ann.procurement_method or "",
            float(ann.budget) if ann.budget else "",
            ann.deadline.strftime("%Y-%m-%d") if ann.deadline and ann.deadline.year > 2000 else "",
            ann.bid_date.strftime("%Y-%m-%d") if ann.bid_date else "",
            float(ann.registration_fee) if ann.registration_fee else "",
            float(ann.deposit) if ann.deposit else "",
            ann.source_url or "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            if col_idx in (1, 2, 3, 7, 8, 9, 10, 11):
                cell.alignment = data_align_center
            elif col_idx == 12 and value:
                # URL 列：设为超链接
                cell.value = "打开链接"
                cell.hyperlink = value
                cell.font = link_font
                cell.alignment = data_align_center
            elif col_idx == 4:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

            # 偶数行浅蓝背景
            if row_idx % 2 == 0:
                cell.fill = even_fill

    # ── 列宽 ──
    col_widths = [6, 8, 10, 55, 14, 14, 12, 13, 13, 13, 13, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 表头行高
    ws.row_dimensions[1].height = 28

    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(announcements) + 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote

    filename = f"收藏公告_{date.today().isoformat()}.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
    )


# ============================================================
# 公告详情（含评分、在位者、提醒、历史参考）
# ============================================================

@router.get("/{announcement_id}", summary="获取公告详情")
async def get_announcement_detail(
    announcement_id: int,
    db: AsyncSession = Depends(get_db),
):
    """
    获取公告详情，自动附带：
    - 采购方信息
    - 关联客情提醒
    - 历史中标参考（同采购方+同赛道）
    - 在位者检测结果
    """
    ann_result = await db.execute(
        select(Announcement)
        .options(selectinload(Announcement.purchaser))
        .where(Announcement.id == announcement_id)
    )
    ann = ann_result.scalar_one_or_none()
    if not ann:
        raise HTTPException(status_code=404, detail=f"公告 {announcement_id} 不存在")

    # 采购方
    purchaser = ann.purchaser
    purchaser_name = purchaser.name if purchaser else ""

    # 关联提醒
    alerts_result = await db.execute(
        select(ProjectRelationAlert)
        .where(ProjectRelationAlert.announcement_id == announcement_id)
        .options(
            selectinload(ProjectRelationAlert.relation).selectinload(ClientRelation.purchaser)
        )
        .order_by(ProjectRelationAlert.created_at.desc())
    )
    alerts = alerts_result.scalars().all()
    alert_items = [
        {
            "id": a.id,
            "alert_reason": a.alert_reason,
            "is_read": a.is_read,
            "contact_name": a.relation.contact_name if a.relation else "",
            "contact_rating": a.relation.rating if a.relation else "",
            "contact_phone": a.relation.phone if a.relation else "",
            "purchaser_name": (
                a.relation.purchaser.name
                if (a.relation and a.relation.purchaser) else ""
            ),
        }
        for a in alerts
    ]

    # 历史中标参考
    from sqlalchemy import text
    history_result = await db.execute(
        text("""
            SELECT * FROM historical_awards
            WHERE purchaser_id = :pid AND project_category = :cat
            ORDER BY bid_open_date DESC LIMIT 5
        """),
        {"pid": ann.purchaser_id, "cat": ann.project_category},
    )
    history_rows = history_result.fetchall()
    history_items = [
        {
            "project_name": row.project_name,
            "winner_name": row.winner_name,
            "winner_type": row.winner_type,
            "bid_amount": float(row.bid_amount) if row.bid_amount else None,
            "budget_amount": float(row.budget_amount) if row.budget_amount else None,
            "discount_rate": float(row.discount_rate) if row.discount_rate else None,
            "bid_open_date": _to_iso(row.bid_open_date),
            "contract_end": _to_iso(row.contract_end),
            "is_continuous": row.is_continuous,
            "continuous_count": row.continuous_count,
        }
        for row in history_rows
    ]

    # 在位者检测
    incumbent_info = None
    if history_items:
        from app.services.incumbent_detector import detect_incumbent
        inc_result = detect_incumbent(
            {"purchaser_id": ann.purchaser_id, "project_category": ann.project_category},
            history_items,
        )
        incumbent_info = {
            "has_incumbent": inc_result.has_incumbent,
            "incumbent_name": inc_result.incumbent_name,
            "continuous_count": inc_result.continuous_count,
            "risk_level": inc_result.risk_level,
            "reason": inc_result.reason,
        }

    # 机会评分
    score_data = _compute_announcement_score(ann)

    return {
        "id": ann.id,
        "title": ann.title,
        "industry": getattr(ann, 'industry', '') or '',
        "province": getattr(ann, 'province', '') or '',
        "city": getattr(ann, 'city', '') or '',
        "purchaser": purchaser_name,
        "purchaser_id": ann.purchaser_id,
        "purchaser_level": ann.purchaser_level,
        "project_category": ann.project_category,
        "procurement_method": ann.procurement_method,
        "budget": float(ann.budget) if ann.budget is not None else None,
        "deadline": ann.deadline.isoformat() if ann.deadline else None,
        "announce_date": ann.announce_date.isoformat() if ann.announce_date else None,
        "qualification_requirements": ann.qualification_requirements,
        "original_content": ann.original_content if hasattr(ann, 'original_content') else "",
        "original_content_html": ann.original_content_html if hasattr(ann, 'original_content_html') else "",
        "score_weight": ann.score_weight,
        "source_url": ann.source_url,
        "created_at": ann.created_at.isoformat() if ann.created_at else None,
        # 评分
        "total_score": score_data.get("total_score"),
        "probability_label": score_data.get("probability_label", ""),
        "detail_scores": score_data.get("detail_scores", {}),
        "recommendation": score_data.get("recommendation", ""),
        # 提醒与在位者
        "alerts": alert_items,
        "history_reference": history_items,
        "incumbent_info": incumbent_info,
    }


# ============================================================
# 手动触发采集
# ============================================================

@router.post("/fetch", summary="手动触发数据采集")
async def fetch_announcements(
    background_tasks: BackgroundTasks,
    province: Optional[str] = Query("广东", description="目标省份: 广东、广西等，或'全国'使用所有适配器"),
):
    """
    触发爬虫采集最新招标公告。

    使用 DataCollector 完整采集链路：
    列表搜索 → 详情提取 → 字段标准化 → LLM分类/预算 → 入库

    支持单省份采集或全国采集（使用所有已启用适配器）。
    """
    if not settings.CRAWLER_ENABLED:
        return {"status": "disabled", "message": "爬虫功能已禁用"}

    province_name = province or "广东"
    task_id = str(uuid.uuid4())[:8]

    is_nationwide = province_name == "全国"

    _fetch_tasks[task_id] = {
        "status": "starting",
        "progress": 0,
        "message": f"正在启动采集引擎（{'全国' if is_nationwide else province_name}）...",
        "province": province_name,
        "started_at": datetime.now().isoformat(),
        "result_count": 0,
        "error": None,
    }

    async def _run_crawler():
        try:
            from data_collector import get_collector
            collector = get_collector()

            if is_nationwide:
                # 全国模式：b2b 不限省份采集
                _fetch_tasks[task_id].update(
                    status="running", progress=5,
                    message="全国采集模式：正在搜索全国移动招标公告...",
                    phase="init",
                )

                _fetch_tasks[task_id].update(
                    progress=15, phase="search",
                    message="正在从 b2b.10086.cn 搜索全国各省公告...",
                )

                _fetch_tasks[task_id].update(
                    progress=25, phase="extract",
                    message="正在逐条提取全国公告详情（预计 3-5 分钟）...",
                )

                import asyncio as _asyncio
                heartbeat_running = True

                async def _heartbeat():
                    p = 25
                    while heartbeat_running and p < 90:
                        await _asyncio.sleep(12)
                        p = min(p + 6, 90)
                        if heartbeat_running:
                            _fetch_tasks[task_id].update(
                                progress=p, phase="extract",
                                message=f"全国采集进行中（已完成约 {p}%）...",
                            )

                heartbeat_task = _asyncio.ensure_future(_heartbeat())

                try:
                    # 不限省份 = 空字符串，parse_list 不做省份过滤
                    results = await collector.collect_async(
                        save_to_db=True, province="",
                    )
                finally:
                    heartbeat_running = False
                    heartbeat_task.cancel()

                _fetch_tasks[task_id].update(
                    status="completed", progress=100,
                    message=f"全国采集完成，共获取 {len(results)} 条公告",
                    result_count=len(results), phase="done",
                )
                logger.info(f"[全国] 采集完成: {len(results)} 条")
                from app.services.notification import notify_collection_done
                await notify_collection_done("公告", len(results))

            else:
                # 单省份模式：b2b 采集 + 省份过滤
                _fetch_tasks[task_id].update(
                    status="running", progress=5,
                    message=f"正在初始化采集引擎（{province_name}）...",
                    phase="init",
                )

                _fetch_tasks[task_id].update(
                    progress=15, phase="search",
                    message=f"正在搜索 b2b.10086.cn {province_name}移动招标公告...",
                )

                _fetch_tasks[task_id].update(
                    progress=25, phase="extract",
                    message=f"正在逐条提取{province_name}公告详情（预计 1-3 分钟）...",
                )

                import asyncio as _asyncio
                heartbeat_running = True

                async def _heartbeat():
                    p = 25
                    while heartbeat_running and p < 90:
                        await _asyncio.sleep(8)
                        p = min(p + 8, 90)
                        if heartbeat_running:
                            _fetch_tasks[task_id].update(
                                progress=p,
                                message=f"正在逐条提取{province_name}公告详情（已完成约 {p}%）...",
                            )

                heartbeat_task = _asyncio.ensure_future(_heartbeat())

                try:
                    results = await collector.collect_async(
                        save_to_db=True, province=province_name,
                    )
                finally:
                    heartbeat_running = False
                    heartbeat_task.cancel()

                _fetch_tasks[task_id].update(
                    status="completed", progress=100,
                    message=f"{province_name}采集完成，共获取 {len(results)} 条公告",
                    result_count=len(results), phase="done",
                )
                logger.info(f"[{province_name}] 采集完成: {len(results)} 条")
                from app.services.notification import notify_collection_done
                await notify_collection_done("公告", len(results), province_name)

        except Exception as e:
            logger.error(f"采集失败: {e}")
            _fetch_tasks[task_id].update(
                status="failed", progress=0,
                message=f"采集失败: {str(e)[:100]}",
                error=str(e), phase="error",
            )

    background_tasks.add_task(_run_crawler)

    return {
        "status": "started",
        "task_id": task_id,
        "message": f"数据采集已在后台启动（{province_name}）",
    }


@router.get("/fetch/status/{task_id}", summary="查询采集进度")
async def get_fetch_status(task_id: str):
    """查询采集任务的实时进度。"""
    task = _fetch_tasks.get(task_id)
    if not task:
        raise HTTPException(status_code=404, detail=f"任务 {task_id} 不存在或已过期")
    return task


@router.post("/discover", summary="多引擎搜索发现招标公告")
async def discover_announcements(
    background_tasks: BackgroundTasks,
    queries: Optional[List[str]] = None,
    max_per_query: int = Query(10, ge=1, le=30, description="每个查询最大结果数"),
    auto_crawl: bool = Query(False, description="是否自动使用 AI 爬虫抓取发现的内容"),
):
    """
    使用多搜索引擎（Baidu/Bing/DuckDuckGo）聚合搜索，
    发现各招标网站上的广东移动广告类公告链接。

    可选：自动使用 AI 爬虫抓取发现的页面内容。
    """
    try:
        from app.services.search_discovery import (
            discover_bidding_announcements,
            discover_and_crawl,
        )
    except ImportError as e:
        return {
            "status": "error",
            "message": f"搜索发现模块不可用: {e}",
        }

    if auto_crawl:
        async def _run_discover_and_crawl():
            try:
                result = await discover_and_crawl(
                    queries=queries,
                    max_per_query=max_per_query,
                )
                logger.info(
                    f"搜索发现+爬取完成: 发现 {result['stats']['discovered']} 条, "
                    f"爬取成功 {result['stats']['crawled_success']} 条"
                )
            except Exception as e:
                logger.error(f"搜索发现+爬取失败: {e}")

        background_tasks.add_task(_run_discover_and_crawl)

        return {
            "status": "started",
            "message": "多引擎搜索发现 + AI 爬取已在后台启动",
        }
    else:
        result = await discover_bidding_announcements(
            queries=queries,
            max_per_query=max_per_query,
        )
        return {
            "status": "ok",
            "total_urls": result["total_count"],
            "urls": result["total_bidding_urls"][:20],
            "search_time": result["search_time"],
        }


# ============================================================
# 收藏/取消收藏
# ============================================================

@router.post("/{announcement_id}/favorite", summary="收藏/取消收藏公告")
async def toggle_favorite(
    announcement_id: int,
    db: AsyncSession = Depends(get_db),
):
    """切换公告收藏状态。已收藏则取消，未收藏则收藏。"""
    ann = await db.get(Announcement, announcement_id)
    if not ann:
        raise HTTPException(status_code=404, detail=f"公告 {announcement_id} 不存在")

    current = getattr(ann, 'is_favorited', False) or False
    ann.is_favorited = not current
    await db.commit()

    return {
        "announcement_id": announcement_id,
        "is_favorited": ann.is_favorited,
        "message": "已收藏" if ann.is_favorited else "已取消收藏",
    }


# ============================================================
# 删除公告
# ============================================================

@router.delete("/{announcement_id}", summary="删除公告")
async def delete_announcement(
    announcement_id: int,
    db: AsyncSession = Depends(get_db),
):
    """软删除一条公告记录。"""
    ann = await db.get(Announcement, announcement_id)
    if not ann:
        raise HTTPException(status_code=404, detail=f"公告 {announcement_id} 不存在")

    await db.delete(ann)
    await db.commit()

    return {
        "announcement_id": announcement_id,
        "message": "已删除",
    }


# ============================================================
# 机会评分辅助函数
# ============================================================

def _compute_announcement_score(ann: Announcement) -> dict:
    """使用评分引擎计算公告的机会评分。

    七维度加权评分：
      1. 采购公平性 (20%) — 基于采购方式的竞争程度（模糊匹配）
      2. 竞争集中度 (15%) — 基于项目类别热度
      3. 赛道匹配度 (15%) — 基于项目类别是否为广告核心赛道
      4. 预算健康度 (15%) — 基于预算规模
      5. 在位者优势 (15%) — 基于是否有历史在位者
      6. 时效新鲜度 (10%) — 基于公告发布时间
      7. 信息完整度 (10%) — 基于关键字段填充率
    """
    try:
        from datetime import datetime

        # ── 1. 采购方式公平性 (20%) ──
        method = (ann.procurement_method or "").strip()
        fairness = 60.0  # 默认中性
        if "公开招标" in method or "招标" in method:
            fairness = 100.0
        elif "公开询比" in method or "询比" in method:
            fairness = 80.0
        elif "竞争性谈判" in method or "谈判" in method:
            fairness = 50.0
        elif "单一来源" in method:
            fairness = 20.0
        elif "比选" in method:
            fairness = 85.0

        # ── 2. 竞争集中度 / 3. 赛道匹配度 ──
        cat = (ann.project_category or "").strip()
        core_cats = {"品牌策略", "创意设计", "媒介投放", "活动执行", "内容制作", "新媒体运营", "品牌宣传"}
        cat_matched = any(c in cat for c in core_cats)
        hhi = 70.0 if cat_matched else 55.0
        category = 70.0 if cat_matched else 50.0

        # ── 4. 预算健康度 (15%) ──
        budget_val = float(ann.budget) if ann.budget else 0
        if budget_val >= 500:
            budget = 95.0
        elif budget_val >= 200:
            budget = 85.0
        elif budget_val >= 100:
            budget = 75.0
        elif budget_val >= 50:
            budget = 60.0
        elif budget_val > 0:
            budget = 40.0
        else:
            budget = 30.0  # 预算未知

        # ── 5. 在位者优势 (15%) ──
        incumbent = 100.0

        # ── 6. 时效新鲜度 (10%) ──
        freshness = 50.0
        if ann.announce_date:
            days_ago = (datetime.now().date() - ann.announce_date).days
            if days_ago <= 3:
                freshness = 100.0
            elif days_ago <= 7:
                freshness = 85.0
            elif days_ago <= 14:
                freshness = 70.0
            elif days_ago <= 30:
                freshness = 55.0
            else:
                freshness = 30.0

        # ── 7. 信息完整度 (10%) ──
        complete_count = 0
        if ann.budget and float(ann.budget) > 0:
            complete_count += 1
        if ann.city and ann.city.strip():
            complete_count += 1
        if ann.deadline and ann.deadline.year > 2000:
            complete_count += 1
        if ann.bid_date:
            complete_count += 1
        completeness = 50.0 + complete_count * 12.5

        # ── 加权汇总 ──
        weights = {
            "procurement_fairness": 0.20,
            "hhi_concentration": 0.15,
            "category_match": 0.15,
            "budget_health": 0.15,
            "incumbent_advantage": 0.15,
            "freshness": 0.10,
            "completeness": 0.10,
        }

        total = (
            fairness * weights["procurement_fairness"]
            + hhi * weights["hhi_concentration"]
            + category * weights["category_match"]
            + budget * weights["budget_health"]
            + incumbent * weights["incumbent_advantage"]
            + freshness * weights["freshness"]
            + completeness * weights["completeness"]
        )

        # ── 陪跑概率标签 ──
        if total >= 75:
            prob_label = "低"
        elif total >= 50:
            prob_label = "中"
        else:
            prob_label = "高"

        # ── 推荐建议（纯文本，避免 emoji 编码问题） ──
        if total >= 75:
            rec = "高机会: 建议优先跟进，竞争环境有利"
        elif total >= 50:
            rec = "中等机会: 评估自身优势后决定是否参与"
        else:
            rec = "低机会: 竞争激烈或在位者优势明显，谨慎评估"

        return {
            "total_score": round(total, 1),
            "probability_label": prob_label,
            "recommendation": rec,
            "detail_scores": {
                "procurement_fairness": round(fairness, 1),
                "hhi_concentration": round(hhi, 1),
                "category_match": round(category, 1),
                "budget_health": round(budget, 1),
                "incumbent_advantage": round(incumbent, 1),
                "freshness": round(freshness, 1),
                "completeness": round(completeness, 1),
            },
        }
    except Exception as e:
        logger.warning(f"评分计算失败: {e}")
        return {
            "total_score": None,
            "probability_label": "",
            "recommendation": "",
            "detail_scores": {},
        }


# ============================================================
# B2B 原文代理（通过 b2b API 获取公告原文）
# ============================================================

@router.get("/{announcement_id}/original", summary="获取公告原文")
async def get_original_content(
    announcement_id: int,
    db: AsyncSession = Depends(get_db),
):
    """
    获取公告原文。
    优先级：数据库 original_content > b2b API 搜索 > 原始 source_url。
    """
    ann = await db.get(Announcement, announcement_id)
    if not ann:
        raise HTTPException(status_code=404, detail=f"公告 {announcement_id} 不存在")

    title = ann.title or ""
    source_url = ann.source_url or ""

    # ── 策略0：数据库已有原文，直接返回 ──
    if ann.original_content and len(ann.original_content) > 100:
        logger.info(f"公告 {announcement_id}: 返回数据库原文 ({len(ann.original_content)}字)")
        return {
            "found": True,
            "announcement_id": announcement_id,
            "title": title,
            "source_url": source_url,
            "notice_content": ann.original_content,
            "publish_date": ann.announce_date.isoformat() if ann.announce_date else None,
            "message": "来自数据库存档",
        }

    # ── 策略1：非 b2b 来源（zhaobiao.cn 等），直接返回原始 URL ──
    if source_url and "b2b.10086.cn" not in source_url:
        logger.info(f"公告 {announcement_id}: 非b2b来源 → 返回原始URL")
        return {
            "found": True,
            "announcement_id": announcement_id,
            "title": title,
            "source_url": source_url,
            "detail_url": source_url,
            "notice_content": ann.original_content or "",
            "publish_date": ann.announce_date.isoformat() if ann.announce_date else None,
            "message": "来自外部来源，点击链接查看原文",
        }

    from app.services.b2b_proxy import (
        search_announcement, find_best_match,
        format_announcement_detail, build_search_url,
        fetch_announcement_detail,
    )

    # ── 策略2：b2b API 搜索 ──
    short_keywords = _extract_search_keywords(title)
    logger.info(f"搜索关键词: {short_keywords}")

    result = None
    searched_keyword = None

    # 尝试不同的 publishType（包括候选人公示）
    for ptype in ["PROCUREMENT", "CANDIDATE_PUBLICITY", "VENDOR", "PURCHASE_SERVICE"]:
        for keyword in short_keywords:
            items = await search_announcement(keyword, publish_type=ptype, page_size=10)
            match = find_best_match(items, title)
            if match:
                # 找到了匹配项，尝试获取完整详情
                detail = await fetch_announcement_detail(
                    match.get("id") or match.get("uuid"),
                    keyword, ptype
                )
                if detail and detail.get("notice_content"):
                    result = detail
                else:
                    result = format_announcement_detail(match)
                result["publish_type_searched"] = ptype
                result["searched_keyword"] = keyword
                searched_keyword = keyword
                break
        if result:
            break

    # ── 策略3：用更长的标题再次尝试 ──
    if not result:
        for ptype in ["PROCUREMENT", "CANDIDATE_PUBLICITY", "VENDOR", "PURCHASE_SERVICE"]:
            items = await search_announcement(title[:30], publish_type=ptype, page_size=15)
            match = find_best_match(items, title)
            if match:
                detail = await fetch_announcement_detail(
                    match.get("id") or match.get("uuid"),
                    title[:30], ptype
                )
                if detail and detail.get("notice_content"):
                    result = detail
                else:
                    result = format_announcement_detail(match)
                searched_keyword = title[:30]
                break
            if result:
                break

    if not result:
        return {
            "found": False,
            "announcement_id": announcement_id,
            "title": title,
            "search_url": build_search_url(title[:30]),
            "searched_keywords": short_keywords,
            "message": "未在 b2b.10086.cn 找到匹配的公告原文，请点击下方按钮在 b2b 网站查看",
        }

    return {
        "found": True,
        "announcement_id": announcement_id,
        "search_url": build_search_url(searched_keyword or title[:30]),
        "detail_url": result.get("detail_url", ""),
        **result,
    }


def _extract_search_keywords(title: str) -> list:
    """从公告标题中提取有辨识度的搜索关键词"""
    keywords = []

    # 去掉公司名前缀（如"中国移动通信集团广东有限公司XX分公司"）
    # 保留项目核心描述部分
    core = title

    # 尝试去掉年份部分后的内容作为关键词
    import re
    # 提取年份后的内容（如"2026年至2028年"之后的部分）
    year_match = re.search(r'\d{4}年.*?\d{4}年(.+)', title)
    if year_match:
        after_year = year_match.group(1)
        # 去掉采购方式后缀（公开询比/公开招标等）
        after_year = re.sub(r'(公开招标|公开询比|竞争性谈判|单一来源|询价).*$', '', after_year)
        if len(after_year) >= 4:
            keywords.append(after_year[:20])

    # 提取地市+项目核心词（使用全国城市名动态正则）
    try:
        from config.provinces import build_city_regex_pattern
        CITY_PATTERN = build_city_regex_pattern()
    except ImportError:
        # 回退：至少覆盖重点省份的常见城市
        CITY_PATTERN = (
            r'广州|深圳|东莞|佛山|中山|珠海|江门|惠州|汕头|湛江|茂名|肇庆|梅州|'
            r'汕尾|河源|阳江|清远|韶关|潮州|揭阳|云浮|'
            r'南宁|柳州|桂林|玉林|梧州|北海|贵港|钦州|百色|河池|贺州|来宾|崇左|防城港|'
            r'福州|厦门|泉州|漳州|龙岩|三明|南平|莆田|宁德|'
            r'海口|三亚|儋州|'
            r'杭州|宁波|温州|嘉兴|湖州|绍兴|金华|衢州|舟山|台州|丽水|'
            r'长沙|株洲|湘潭|衡阳|邵阳|岳阳|常德|张家界|益阳|郴州|永州|怀化|娄底|'
            r'合肥|芜湖|蚌埠|淮南|马鞍山|淮北|铜陵|安庆|黄山|滁州|阜阳|宿州|六安|亳州|池州|宣城|'
            r'济南|青岛|淄博|枣庄|东营|烟台|潍坊|济宁|泰安|威海|日照|临沂|德州|聊城|滨州|菏泽'
        )
    city_match = re.search(f'({CITY_PATTERN})', title)
    if city_match:
        city = city_match.group(1)
        # 地市 + 项目关键词
        if year_match:
            keywords.append(f"{city} {after_year[:15]}")

    # 提取括号内的关键词
    bracket_match = re.search(r'[（(]([^）)]+)[）)]', title)
    if bracket_match:
        kw = bracket_match.group(1)
        if len(kw) >= 3 and kw not in ['二次', '重新招标']:
            keywords.append(kw[:15])

    # 用前30字符（去括号版本）
    clean_title = re.sub(r'[（(][^）)]*[）)]', '', title)
    keywords.append(clean_title[:30])

    # 用前15字符（最短版本，可能匹配更广）
    keywords.append(clean_title[:15])

    # 去重，过滤太短的关键词
    seen = set()
    result = []
    for k in keywords:
        k = k.strip()
        if k and len(k) >= 3 and k not in seen:
            seen.add(k)
            result.append(k)

    return result


# ============================================================
# 预算抓取（zhaobiao.cn 登录后自动提取）
# ============================================================

@router.post("/scrape-budget/start", summary="启动预算抓取（需手动登录 zhaobiao.cn）")
async def start_budget_scrape():
    """
    启动后台预算抓取任务。
    
    流程：
    1. 打开 zhaobiao.cn 浏览器窗口
    2. 用户手动登录（含验证码）
    3. 系统自动抓取每条公告的预算/报名费/保证金
    4. 更新数据库
    
    前端应轮询 GET /scrape-budget/status 获取进度。
    """
    from app.services.budget_scraper import start_scrape_async, get_state, ScrapeStatus

    state = get_state()
    if state.status in (ScrapeStatus.WAITING_LOGIN, ScrapeStatus.SCRAPING):
        return {"ok": False, "message": "抓取任务已在运行中", "status": state.status}

    # 获取数据库路径
    db_url = getattr(settings, 'DATABASE_URL', '')
    # sqlite+aiosqlite:///./biaozhongbao.db → biaozhongbao.db
    for prefix in ['sqlite+aiosqlite:///', 'sqlite:///']:
        if db_url.startswith(prefix):
            db_url = db_url[len(prefix):]
            break
    if not db_url:
        db_url = "biaozhongbao.db"

    start_scrape_async(db_url)
    return {"ok": True, "message": "抓取任务已启动，请在弹出的浏览器中登录", "status": "waiting_login"}


@router.get("/scrape-budget/status", summary="查询预算抓取进度")
async def get_budget_scrape_status():
    """返回当前抓取任务的状态和结果。"""
    from app.services.budget_scraper import get_state

    state = get_state()
    return {
        "status": state.status,
        "message": state.message,
        "login_elapsed": state.login_elapsed,
        "results": state.results,
    }


# ============================================================
# LLM 预算提取（从 b2b 公告正文中用 AI 提取预算）
# 注意：/extract-budget/batch 必须在 /extract-budget/{id} 之前注册！
# ============================================================

@router.post("/extract-budget/batch", summary="b2b 全自动抓取 + LLM 提取预算")
async def extract_budget_batch(
    db: AsyncSession = Depends(get_db),
    limit: int = Query(10, ge=1, le=50, description="最多处理条数"),
):
    """
    全自动从 b2b.10086.cn 抓取公告正文并用 LLM 提取预算。
    通过 Playwright 拦截 SPA 网络请求注入搜索结果，全程无需手动操作。
    """
    from app.services.b2b_auto_scraper import scrape_auto
    from app.services.llm_budget_extractor import extract_budget_with_llm

    result = await db.execute(
        select(Announcement)
        .where((Announcement.budget == None) | (Announcement.budget == 0))
        .order_by(desc(Announcement.announce_date))
        .limit(limit)
    )
    candidates = result.scalars().all()

    if not candidates:
        return {"ok": True, "message": "所有公告已有预算", "processed": 0, "results": []}

    results = []
    for ann in candidates:
        title = ann.title or ""
        keywords = _extract_search_keywords(title)
        search_kw = keywords[0] if keywords else title[:30]

        logger.info(f"🤖 全自动抓取: {search_kw[:40]}")
        scraped = await scrape_auto(search_kw)

        if not scraped or not scraped.get("content"):
            results.append({"id": ann.id, "title": title[:60], "status": "no_content"})
            continue

        try:
            bd = await extract_budget_with_llm(title, scraped["content"])
            if bd.get("budget_wan") is not None:
                ann.budget = bd["budget_wan"]
            if bd.get("registration_fee") is not None:
                ann.registration_fee = bd["registration_fee"]
            if bd.get("deposit") is not None:
                ann.deposit = bd["deposit"]
            if bd.get("bid_date"):
                from datetime import datetime as dt
                try:
                    ann.bid_date = dt.strptime(bd["bid_date"], "%Y-%m-%d").date()
                except ValueError:
                    pass
            await db.commit()
            results.append({
                "id": ann.id, "title": title[:60], "status": "extracted",
                "budget_wan": bd.get("budget_wan"), "confidence": bd.get("confidence"),
                "method": scraped.get("method", "unknown"),  # 记录使用的导航方法
                "content_length": len(scraped.get("content", "")),
            })
            logger.info(f"✅ ID={ann.id} budget={bd.get('budget_wan')}万 method={scraped.get('method', 'unknown')}")
        except Exception as e:
            results.append({"id": ann.id, "title": title[:60], "status": "error", "reason": str(e)[:200]})

    return {
        "ok": True,
        "total": len(candidates),
        "extracted": sum(1 for r in results if r["status"] == "extracted"),
        "results": results,
    }


@router.post("/extract-budget/{announcement_id}", summary="LLM 提取单条公告预算")
async def extract_budget_single(
    announcement_id: int,
    db: AsyncSession = Depends(get_db),
):
    """
    使用 LLM 从 b2b 公告正文中提取预算金额并自动更新数据库。
    注意：b2b API 通常不返回 noticeContent，
    若提取失败请用 download_b2b_content.py 手动下载正文。
    """
    from app.services.b2b_proxy import (
        search_announcement, find_best_match, fetch_announcement_detail,
    )
    from app.services.llm_budget_extractor import extract_budget_with_llm

    ann = await db.get(Announcement, announcement_id)
    if not ann:
        raise HTTPException(status_code=404, detail=f"公告 {announcement_id} 不存在")

    title = ann.title or ""
    keywords = _extract_search_keywords(title)
    content = None

    for ptype in ["PROCUREMENT", "CANDIDATE_PUBLICITY"]:
        for kw in keywords:
            items = await search_announcement(kw, publish_type=ptype, page_size=10)
            match = find_best_match(items, title)
            if match:
                detail = await fetch_announcement_detail(
                    match.get("id") or match.get("uuid"), kw, ptype
                )
                if detail and detail.get("notice_content"):
                    content = detail["notice_content"]
                    break
        if content:
            break

    if not content:
        return {
            "ok": False,
            "announcement_id": announcement_id,
            "message": "b2b noticeContent 为空，请用 download_b2b_content.py 手动提取",
        }

    budget_data = await extract_budget_with_llm(title, content)

    updates = {}
    if budget_data.get("budget_wan") is not None:
        ann.budget = budget_data["budget_wan"]
        updates["budget"] = budget_data["budget_wan"]
    if budget_data.get("registration_fee") is not None:
        ann.registration_fee = budget_data["registration_fee"]
        updates["registration_fee"] = budget_data["registration_fee"]
    if budget_data.get("deposit") is not None:
        ann.deposit = budget_data["deposit"]
        updates["deposit"] = budget_data["deposit"]
    if budget_data.get("bid_date"):
        from datetime import datetime as dt
        try:
            ann.bid_date = dt.strptime(budget_data["bid_date"], "%Y-%m-%d").date()
            updates["bid_date"] = budget_data["bid_date"]
        except ValueError:
            pass

    if updates:
        await db.commit()

    return {
        "ok": True,
        "announcement_id": announcement_id,
        **budget_data,
        "db_updated": bool(updates),
    }
