#!/usr/bin/env python3
"""
三大运营商广告类招标公告 — 每日爬取 + LLM 精判 + Excel 输出

流程:
  1. 初始化三个运营商适配器（移动/电信/联通）
  2. 逐平台搜索广告关键词，获取候选公告列表
  3. 关键词粗筛（使用 SAFETY_KEYWORDS）
  4. LLM 精判 — 确认广告属性 + 归类到 8 个赛道
  5. 去重合并 → 输出格式化 Excel

Usage:
    cd D:\\bzb\\backend
    python scripts/scrape_ad_bidding_daily.py

Output:   D:\\bzb\\output\\三大运营商广告类招标公告_yyyy-MM-dd.xlsx
          C:\\Users\\18826\\Desktop\\三大运营商广告类招标公告_yyyy-MM-dd.xlsx
"""

import json
import logging
import os
import re
import sys
import time
from datetime import datetime, date
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path

# ── 路径设置 ──
BACKEND_DIR = Path(__file__).resolve().parent.parent  # D:\bzb\backend
sys.path.insert(0, str(BACKEND_DIR))

# ── 加载 .env ──
try:
    from dotenv import load_dotenv
    _env_path = BACKEND_DIR / ".env"
    if _env_path.exists():
        load_dotenv(_env_path)
except ImportError:
    pass

# ── 日志 ──
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("scrape_ad_bidding")

# ============================================================
# 0. 配置
# ============================================================

LLM_API_BASE = os.getenv("BZB_LLM_API_BASE", "https://api.deepseek.com/v1")
LLM_API_KEY = os.getenv("BZB_LLM_API_KEY", os.getenv("DEEPSEEK_API_KEY", ""))
LLM_MODEL = os.getenv("BZB_LLM_MODEL", "deepseek-chat")

OUTPUT_DIR = BACKEND_DIR.parent / "output"  # D:\bzb\output
DESKTOP_DIR = Path("C:/Users/18826/Desktop")

# ============================================================
# 1. 广告赛道关键词（8 个赛道）
# ============================================================

# ── 安全词（命中即视为广告候选）──
SAFETY_KEYWORDS: List[str] = [
    # 广告创意设计
    "广告设计", "广告创意", "广告策划", "广告全案", "广告全案策划",
    "广告整合推广", "广告代理", "平面设计", "视觉设计", "创意设计",
    "VI设计", "VI标识", "品牌VI", "视觉识别系统", "形象设计", "品牌设计",
    "全案策划", "品牌形象设计", "品牌视觉", "视觉创意",
    # 物料制作印刷
    "广告物料", "广告物料制作", "广告制作", "宣传物料", "宣传物料制作",
    "物料制作", "物料采购", "喷绘制作", "写真制作", "标识标牌", "门头招牌",
    "灯箱广告", "发光字", "户外广告牌", "营业厅门头", "营业厅背景墙",
    "广告宣传品", "印刷品", "宣传品印制", "广告宣传物料", "营销物料",
    "促销物料", "展架", "易拉宝", "横幅", "宣传单页", "宣传册", "画册",
    "海报", "展板",
    # 活动策划执行
    "活动策划", "活动执行", "活动服务", "营销活动", "市场活动", "推广活动",
    "品牌活动", "客户活动", "客户服务活动", "客户关怀", "集团客户活动",
    "政企客户活动", "校园营销", "校园活动", "社区活动", "发布会", "推介会",
    "展会", "展览展示", "会议会展", "路演", "快闪", "运动会", "文体活动",
    "工会活动", "党建活动", "文化宣传活动", "促销活动", "庆典活动", "年会", "启动仪式",
    # 品牌宣传传播
    "品牌宣传", "品牌推广", "品牌传播", "企业文化宣传", "党建宣传", "新闻宣传",
    "宣传推广", "宣传服务", "品牌建设", "品牌形象", "品牌策划",
    "整合营销", "社会化营销", "新媒体传播", "公关传播", "媒体宣传",
    # 视频内容制作
    "视频制作", "宣传片拍摄", "宣传片制作", "短视频制作", "视频拍摄", "视频剪辑",
    "品牌视频", "形象宣传片", "产品视频", "动画制作", "微电影", "直播",
    # 新媒体运营
    "新媒体运营", "公众号运营", "抖音运营", "视频号运营", "新媒体策划",
    "新媒体内容制作", "新媒体代运营", "内容运营", "H5制作", "直播运营",
    "自媒体运营", "社交媒体运营",
    # 媒介资源投放
    "户外媒介", "户外广告", "社区广告", "社区道闸", "社区电梯广告",
    "公交车身广告", "公交候车亭广告", "地铁广告", "地铁灯箱",
    "高铁站广告", "机场广告", "LED大屏", "户外大牌",
    "电梯框架广告", "电梯视频广告", "出租车广告",
    "网络广告", "互联网广告", "信息流广告", "KOL投放", "达人合作",
    # 渠道营销推广
    "渠道营销", "网格营销", "地推", "促销布展", "门店推广", "厅店推广",
    "营业厅推广", "商圈推广", "社区推广", "校园推广", "政企推广",
    "客户拓展", "市场拓展", "营销支撑", "运营支撑",
    # 广义兜底
    "广告", "宣传", "活动", "品牌", "新媒体", "视频", "物料", "设计",
    "营销", "推广",
]

CATEGORY_RULES: List[Dict[str, Any]] = [
    {"category": "广告创意设计", "keywords": [
        "广告设计", "广告创意", "平面设计", "视觉设计", "创意设计",
        "VI设计", "VI标识", "品牌VI", "视觉识别系统", "形象设计", "品牌设计",
        "广告全案", "全案策划", "广告策划", "广告整合推广",
    ]},
    {"category": "物料制作印刷", "keywords": [
        "广告物料", "广告物料制作", "宣传物料", "宣传物料制作",
        "物料制作", "喷绘制作", "写真制作", "标识标牌",
        "门头招牌", "灯箱", "发光字", "户外广告牌",
        "广告宣传品", "印刷品", "宣传品印制",
        "展架", "易拉宝", "横幅", "宣传单页", "宣传册", "画册", "海报", "展板",
    ]},
    {"category": "活动策划执行", "keywords": [
        "活动策划", "活动执行", "营销活动", "市场活动", "推广活动",
        "品牌活动", "客户活动", "客户服务活动", "客户关怀",
        "集团客户活动", "政企客户活动", "校园营销", "发布会",
        "推介会", "展会", "展览展示", "会议会展", "路演", "快闪",
        "运动会", "文体活动", "工会活动", "党建活动", "促销活动", "年会", "启动仪式",
    ]},
    {"category": "品牌宣传传播", "keywords": [
        "品牌宣传", "品牌推广", "品牌传播", "企业文化宣传",
        "党建宣传", "新闻宣传", "宣传推广", "宣传服务",
        "品牌建设", "品牌形象", "品牌策划",
        "整合营销", "社会化营销", "新媒体传播", "公关传播", "媒体宣传",
    ]},
    {"category": "视频内容制作", "keywords": [
        "视频制作", "宣传片拍摄", "宣传片制作", "短视频制作",
        "视频拍摄", "视频剪辑", "品牌视频", "形象宣传片",
        "产品视频", "动画制作", "微电影", "直播",
    ]},
    {"category": "新媒体运营", "keywords": [
        "新媒体运营", "公众号运营", "抖音运营", "视频号运营",
        "新媒体策划", "新媒体内容制作", "新媒体代运营",
        "内容运营", "H5制作", "直播运营",
    ]},
    {"category": "媒介资源投放", "keywords": [
        "户外媒介", "户外广告", "社区广告", "社区道闸",
        "社区电梯广告", "公交车身广告", "公交候车亭广告",
        "地铁广告", "地铁灯箱", "LED大屏", "户外大牌",
        "电梯框架广告", "电梯视频广告", "出租车广告",
        "网络广告", "互联网广告", "信息流广告", "KOL投放",
    ]},
    {"category": "渠道营销推广", "keywords": [
        "渠道营销", "网格营销", "地推", "促销布展",
        "门店推广", "厅店推广", "营业厅推广", "商圈推广",
        "社区推广", "校园推广", "政企推广",
        "客户拓展", "市场拓展", "营销支撑", "运营支撑",
    ]},
]

# 硬排除词
HARD_EXCLUDE = [
    "基站建设", "光缆铺设", "软件开发", "系统编码", "服务器采购",
    "网络设备", "交换机", "路由器", "机房", "存储设备", "磁盘阵列",
    "服务器", "数据库", "物业管理", "食堂承包", "保安服务", "保洁服务",
    "绿化养护", "消防维保", "消防系统", "安防监控", "空调维修",
    "空调维保", "电梯维保", "电力工程", "土建", "建筑施工",
    "装修工程", "装饰工程", "弱电工程", "强电", "防水工程", "钢结构",
    "车辆维修", "车辆保养", "车辆采购", "运输服务",
    "IT服务", "信息系统", "网络安全", "数据安全",
    "云计算", "大数据", "AI服务", "人工智能",
    "审计服务", "法律服务", "财务咨询", "资产评估",
    "办公用品", "办公设备", "办公家具", "饮用水",
    "桶装水", "劳保用品", "工作服", "制服采购",
    "通信工程", "网络优化", "网络维护", "线路维护",
    "综合配套", "配套施工", "管线工程",
    "设计院", "工程勘察", "勘察设计", "工程监理",
    "招标代理", "造价咨询",
]


def _match_keywords(text: str, keywords: List[str]) -> List[str]:
    """在文本中匹配关键词，返回命中的关键词列表（去重）。"""
    if not text:
        return []
    text_lower = text.lower()
    matched = []
    seen = set()
    for kw in keywords:
        if kw.lower() in text_lower and kw not in seen:
            matched.append(kw)
            seen.add(kw)
    return matched


def _identify_category(title: str, content: str = "") -> str:
    """按 CATEGORY_RULES 顺序匹配赛道。"""
    combined = f"{title} {content}"
    for rule in CATEGORY_RULES:
        if _match_keywords(combined, rule["keywords"]):
            return rule["category"]
    return "其他营销类"


def keyword_filter(title: str, content: str = "") -> Dict[str, Any]:
    """
    关键词粗筛 — 判断是否属于广告类候选。
    返回: {"is_ad_candidate": bool, "category": str, "reason": str}
    """
    combined = f"{title} {content}"

    # 第1层：安全词 → 直接候选
    safeties = _match_keywords(combined, SAFETY_KEYWORDS)
    if safeties:
        return {
            "is_ad_candidate": True,
            "category": _identify_category(title, content),
            "reason": f"命中安全词: {','.join(safeties[:3])}",
        }

    # 第2层：硬排除 → 明确非广告
    excluded = _match_keywords(combined, HARD_EXCLUDE)
    if excluded:
        return {
            "is_ad_candidate": False,
            "category": "",
            "reason": f"命中排除词: {','.join(excluded[:2])}",
        }

    # 非广告业务词也排除
    biz_exclude = ["采购", "招标", "比选"]
    if not _match_keywords(combined, biz_exclude):
        return {
            "is_ad_candidate": False,
            "category": "",
            "reason": "非招标采购类项目",
        }

    return {
        "is_ad_candidate": False,
        "category": "",
        "reason": "未命中广告关键词",
    }


# ============================================================
# 2. LLM 精判
# ============================================================

AD_CATEGORIES_DESC = """
- 广告创意设计：广告设计、VI设计、品牌视觉、全案策划、平面设计
- 物料制作印刷：宣传物料、喷绘写真、标识标牌、门头招牌、印刷品、海报展板
- 活动策划执行：活动策划、路演、发布会、展会、客户活动、校园营销、促销活动
- 品牌宣传传播：品牌推广、整合营销、公关传播、媒体宣传、企业文化建设
- 视频内容制作：宣传片拍摄、视频制作、动画制作、微电影、短视频
- 新媒体运营：公众号运营、抖音运营、直播运营、H5制作、社交媒体运营
- 媒介资源投放：户外广告、社区广告、公交广告、地铁广告、信息流广告、LED大屏
- 渠道营销推广：网格营销、门店推广、地推、商圈推广、客户拓展、营业厅推广
"""

SYSTEM_PROMPT = f"""你是一名运营商招标项目分类专家。请判断招标公告是否属于"广告营销类"。

广告营销类包括以下赛道：
{AD_CATEGORIES_DESC}

不属于广告营销类的示例（应判定为非广告类）：
- 基站建设、光缆铺设、机房设备、铁塔维护
- 软件开发、系统集成、IT运维
- 物业管理、食堂承包、保安保洁
- 空调消防、电力电源、综合布线
- 通信设备采购、网络技术支撑、工程设计勘察

请严格按以下 JSON 格式回复，不要包含其他内容：
{{"is_ad": true/false, "category": "赛道名称或空字符串", "reason": "一句话判断理由(不超过30字)"}}"""


def _build_user_prompt(title: str) -> str:
    return f"请判断以下招标公告是否属于广告营销类：\n\n【项目名称】\n{title}"


def _parse_llm_json(text: str) -> Dict[str, Any]:
    """解析 LLM 返回的 JSON。"""
    try:
        data = json.loads(text)
        return {
            "is_ad": bool(data.get("is_ad", False)),
            "category": str(data.get("category", "")),
            "reason": str(data.get("reason", ""))[:50],
        }
    except json.JSONDecodeError:
        pass
    m = re.search(r'\{[^{}]*"is_ad"[^{}]*\}', text, re.DOTALL)
    if m:
        try:
            data = json.loads(m.group())
            return {
                "is_ad": bool(data.get("is_ad", False)),
                "category": str(data.get("category", "")),
                "reason": str(data.get("reason", ""))[:50],
            }
        except json.JSONDecodeError:
            pass
    m2 = re.search(r'\{[^{}]*"is_advertising"[^{}]*\}', text, re.DOTALL)
    if m2:
        try:
            data = json.loads(m2.group())
            return {
                "is_ad": bool(data.get("is_advertising", False)),
                "category": str(data.get("category", "")),
                "reason": str(data.get("reason", ""))[:50],
            }
        except json.JSONDecodeError:
            pass
    logger.warning(f"LLM响应无法解析为JSON: {text[:100]}")
    return {"is_ad": False, "category": "", "reason": "解析失败"}


def llm_classify(title: str) -> Dict[str, Any]:
    """
    使用 LLM 精判公告是否属于广告营销类。
    返回: {"is_ad": bool, "category": str, "reason": str}
    """
    if not LLM_API_KEY:
        logger.debug("LLM API Key 未配置，跳过精判")
        return {"is_ad": False, "category": "", "reason": "LLM未配置"}

    import httpx
    try:
        with httpx.Client(timeout=20) as client:
            resp = client.post(
                f"{LLM_API_BASE}/chat/completions",
                headers={
                    "Authorization": f"Bearer {LLM_API_KEY}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": LLM_MODEL,
                    "messages": [
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user", "content": _build_user_prompt(title)},
                    ],
                    "temperature": 0.1,
                    "max_tokens": 200,
                },
            )
            resp.raise_for_status()
            data = resp.json()
            result_text = data["choices"][0]["message"]["content"]
    except Exception as e:
        logger.warning(f"LLM 分类调用失败 {title[:40]}: {e}")
        return {"is_ad": True, "category": "", "reason": "LLM调用失败,保留关键字判定"}  # 保底：保留关键词判定

    result = _parse_llm_json(result_text)
    logger.info(f"LLM[{title[:30]}...] → is_ad={result['is_ad']} cat={result['category']}")
    return result


# ============================================================
# 3. 三大平台爬取
# ============================================================

def scrape_b2b_10086() -> List[Dict[str, Any]]:
    """爬取中国移动 B2B 平台 (b2b.10086.cn)"""
    logger.info("=" * 50)
    logger.info("开始爬取：中国移动 B2B")
    logger.info("=" * 50)

    import httpx
    import ssl

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    ctx.options |= 0x4

    results = []
    seen_ids = set()

    LIST_API = "https://b2b.10086.cn/api-b2b/api-sync-es/white_list_api/b2b/publish/queryList"
    keywords = ["广告", "宣传", "品牌", "活动策划", "新媒体", "视频制作", "营销", "设计", "物料", "推广"]
    publish_types = ["PROCUREMENT", "PURCHASE_SERVICE"]

    try:
        with httpx.Client(
            transport=httpx.HTTPTransport(verify=ctx),
            timeout=httpx.Timeout(30),
            headers={
                "Content-Type": "application/json",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            },
        ) as client:
            for kw in keywords:
                for ptype in publish_types:
                    try:
                        time.sleep(2 + 2 * (len(results) % 3))  # 随机延迟
                        resp = client.post(
                            LIST_API,
                            json={
                                "name": kw,
                                "publishType": ptype,
                                "size": 20,
                                "current": 1,
                                "sfactApplColumn5": "PC",
                            },
                        )
                        if resp.status_code == 200:
                            items = resp.json().get("data", {}).get("content", [])
                            for item in items:
                                pid = str(item.get("id", ""))
                                if pid and pid not in seen_ids:
                                    seen_ids.add(pid)
                                    title = item.get("name", "")
                                    if not title:
                                        continue
                                    detail_url = (
                                        f"https://b2b.10086.cn/#/noticeDetail"
                                        f"?publishId={pid}"
                                        f"&publishUuid={item.get('uuid', '')}"
                                        f"&publishType=PROCUREMENT"
                                        f"&publishOneType={item.get('publishOneType', 'PROCUREMENT')}"
                                    )
                                    results.append({
                                        "title": title,
                                        "publish_date": item.get("publishDate", ""),
                                        "detail_url": detail_url,
                                        "notice_type": item.get("publishOneType", ""),
                                        "source": "移动",
                                        "province": item.get("provinceName", ""),
                                        "_raw": item,
                                    })
                    except Exception as e:
                        logger.debug(f"移动搜索 '{kw}' ({ptype}): {e}")
    except Exception as e:
        logger.error(f"移动平台连接失败: {e}")

    logger.info(f"移动平台获取到 {len(results)} 条原始候选")
    return results


def scrape_telecom() -> List[Dict[str, Any]]:
    """爬取中国电信采购平台 (caigou.chinatelecom.com.cn)"""
    logger.info("=" * 50)
    logger.info("开始爬取：中国电信")
    logger.info("=" * 50)

    import httpx
    import ssl

    results = []
    seen_ids = set()

    BASE_URL = "https://caigou.chinatelecom.com.cn"
    LIST_API = "https://caigou.chinatelecom.com.cn/portal/base/announcementJoin/queryListNew"
    keywords = [
        "广告设计", "品牌推广", "活动策划", "新媒体运营",
        "媒介投放", "视频制作", "创意设计", "品牌宣传",
        "广告", "宣传", "活动", "品牌", "新媒体", "视频",
    ]
    ann_types = {"xi9s": "采购公告", "n0eves": "结果公告"}

    try:
        with httpx.Client(
            transport=httpx.HTTPTransport(verify=False),
            timeout=httpx.Timeout(30),
            follow_redirects=True,
            headers={
                "Content-Type": "application/json;charset=UTF-8",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Referer": f"{BASE_URL}/",
                "Origin": BASE_URL,
            },
        ) as client:
            # 获取 Cookie
            try:
                client.get(BASE_URL + "/", follow_redirects=True)
            except Exception:
                pass

            for kw in keywords:
                for ann_type in ann_types:
                    try:
                        time.sleep(1 + len(results) % 3)
                        body = {
                            "pageNum": 1,
                            "pageSize": 20,
                            "type": ann_type,
                            "name": kw,
                        }
                        resp = client.post(LIST_API, json=body)
                        if resp.status_code == 200:
                            data = resp.json()
                            if data.get("code") == 200:
                                items = data.get("data", {}).get("pageInfo", {}).get("list", [])
                                for item in items:
                                    item_id = str(item.get("id", ""))
                                    if item_id and item_id not in seen_ids:
                                        seen_ids.add(item_id)
                                        title = item.get("docTitle", "")
                                        if not title:
                                            continue
                                        detail_url = (
                                            f"{BASE_URL}/portal/system"
                                            f"?page=announcementDetail&type={ann_type}"
                                            f"&announcementId={item_id}"
                                        )
                                        results.append({
                                            "title": title,
                                            "publish_date": item.get("createDate", ""),
                                            "detail_url": detail_url,
                                            "notice_type": ann_types[ann_type],
                                            "source": "电信",
                                            "province": item.get("provinceName", ""),
                                            "_raw": item,
                                        })
                    except Exception as e:
                        logger.debug(f"电信搜索 '{kw}' ({ann_type}): {e}")
    except Exception as e:
        logger.error(f"电信平台连接失败: {e}")

    logger.info(f"电信平台获取到 {len(results)} 条原始候选")
    return results


def scrape_unicom() -> List[Dict[str, Any]]:
    """爬取中国联通采购平台 (chinaunicombidding.cn)"""
    logger.info("=" * 50)
    logger.info("开始爬取：中国联通")
    logger.info("=" * 50)

    import httpx

    results = []
    seen_ids = set()

    BASE_URL = "https://www.chinaunicombidding.cn"
    LIST_API = "https://www.chinaunicombidding.cn/api/v1/bizAnno/getAnnoList"
    keywords = [
        "广告设计", "品牌推广", "活动策划", "新媒体运营",
        "媒介投放", "视频制作", "创意设计", "品牌宣传",
        "广告", "宣传", "活动", "品牌", "新媒体", "视频", "物料", "设计",
        "营销", "推广",
    ]

    try:
        with httpx.Client(
            timeout=httpx.Timeout(30),
            follow_redirects=True,
            headers={
                "Content-Type": "application/json;charset=UTF-8",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Referer": f"{BASE_URL}/",
                "Origin": BASE_URL,
            },
        ) as client:
            # 获取 Cookie
            try:
                client.get(BASE_URL + "/", follow_redirects=True)
            except Exception:
                pass

            for kw in keywords:
                try:
                    time.sleep(1 + len(results) % 3)
                    body = {
                        "pageNo": 1,
                        "pageSize": 10,
                        "modeNo": "BizAnnoVoMtable",
                        "annoName": kw,
                    }
                    resp = client.post(LIST_API, json=body)
                    if resp.status_code == 200:
                        data = resp.json()
                        items = data.get("data", {})
                        if isinstance(items, dict):
                            items = items.get("records", [])
                        if not isinstance(items, list):
                            items = []
                        for item in items:
                            item_id = str(item.get("id", ""))
                            if item_id and item_id not in seen_ids:
                                seen_ids.add(item_id)
                                title = item.get("annoName", "")
                                if not title:
                                    continue
                                detail_url = f"{BASE_URL}/detail?id={item_id}"
                                results.append({
                                    "title": title,
                                    "publish_date": item.get("annoPublicTime", ""),
                                    "detail_url": detail_url,
                                    "notice_type": item.get("annoTypeName", ""),
                                    "source": "联通",
                                    "province": item.get("provinceName", ""),
                                    "_raw": item,
                                })
                except Exception as e:
                    logger.debug(f"联通搜索 '{kw}': {e}")
    except Exception as e:
        logger.error(f"联通平台连接失败: {e}")

    logger.info(f"联通平台获取到 {len(results)} 条原始候选")
    return results


# ============================================================
# 4. 省份/城市提取
# ============================================================

ALL_PROVINCES = [
    "北京", "天津", "上海", "重庆",
    "河北", "山西", "辽宁", "吉林", "黑龙江",
    "江苏", "浙江", "安徽", "福建", "江西", "山东",
    "河南", "湖北", "湖南",
    "广东", "广西", "海南",
    "四川", "贵州", "云南", "西藏",
    "陕西", "甘肃", "青海", "宁夏", "新疆",
    "内蒙古",
    "香港", "澳门", "台湾",
]

ALL_CITIES = [
    "广州", "深圳", "东莞", "佛山", "惠州", "珠海", "中山", "江门",
    "茂名", "揭阳", "汕头", "湛江", "肇庆", "梅州", "汕尾", "河源",
    "阳江", "清远", "韶关", "潮州", "云浮",
    "南京", "无锡", "徐州", "常州", "苏州", "南通", "连云港",
    "淮安", "盐城", "扬州", "镇江", "泰州", "宿迁",
    "杭州", "宁波", "温州", "嘉兴", "湖州", "绍兴", "金华",
    "衢州", "舟山", "台州", "丽水",
    "成都", "绵阳", "德阳", "宜宾", "南充", "泸州", "乐山",
    "武汉", "襄阳", "宜昌", "荆州", "黄石",
    "济南", "青岛", "烟台", "潍坊", "临沂", "淄博",
    "西安", "宝鸡", "咸阳", "渭南",
    "长沙", "株洲", "湘潭", "衡阳", "岳阳", "常德",
    "郑州", "洛阳", "南阳", "新乡", "开封",
    "沈阳", "大连", "鞍山", "抚顺",
    "哈尔滨", "大庆", "齐齐哈尔",
    "福州", "厦门", "泉州", "漳州", "龙岩",
    "昆明", "曲靖", "大理", "丽江",
    "合肥", "芜湖", "蚌埠", "马鞍山",
    "南宁", "柳州", "桂林",
    "石家庄", "唐山", "保定", "邯郸",
    "太原", "大同", "长治",
    "兰州", "天水",
    "呼和浩特", "包头", "鄂尔多斯",
    "贵阳", "遵义",
    "南昌", "九江",
    "海口", "三亚",
    "西宁", "银川", "乌鲁木齐", "拉萨",
]


def extract_province(title: str, province_field: str = "") -> Tuple[str, str]:
    """从标题 + 平台字段提取省份和城市。"""
    # 先测 platform 返回的省份
    if province_field:
        for prov in ALL_PROVINCES:
            if prov in province_field:
                # 从剩余内容找城市
                for city in ALL_CITIES:
                    if city in title:
                        return prov, city
                return prov, ""

    # 从标题提取
    found_prov = ""
    found_city = ""
    for city in ALL_CITIES:
        if city in title:
            found_city = city
            break

    for prov in ALL_PROVINCES:
        if prov in title:
            found_prov = prov
            break

    # 如果城市确定了但省份没确定，从城市倒推省份
    if found_city and not found_prov:
        province_city_map = {
            "广东": ["广州", "深圳", "东莞", "佛山", "惠州", "珠海", "中山", "江门",
                    "茂名", "揭阳", "汕头", "湛江", "肇庆", "梅州", "汕尾", "河源",
                    "阳江", "清远", "韶关", "潮州", "云浮"],
            "江苏": ["南京", "无锡", "徐州", "常州", "苏州", "南通", "连云港",
                    "淮安", "盐城", "扬州", "镇江", "泰州", "宿迁"],
            "浙江": ["杭州", "宁波", "温州", "嘉兴", "湖州", "绍兴", "金华",
                    "衢州", "舟山", "台州", "丽水"],
            "四川": ["成都", "绵阳", "德阳", "宜宾", "南充", "泸州", "乐山"],
            "湖北": ["武汉", "襄阳", "宜昌", "荆州", "黄石"],
            "山东": ["济南", "青岛", "烟台", "潍坊", "临沂", "淄博"],
            "陕西": ["西安", "宝鸡", "咸阳", "渭南"],
            "湖南": ["长沙", "株洲", "湘潭", "衡阳", "岳阳", "常德"],
            "河南": ["郑州", "洛阳", "南阳", "新乡", "开封"],
            "辽宁": ["沈阳", "大连", "鞍山", "抚顺"],
            "黑龙江": ["哈尔滨", "大庆", "齐齐哈尔"],
            "福建": ["福州", "厦门", "泉州", "漳州", "龙岩"],
            "云南": ["昆明", "曲靖", "大理", "丽江"],
            "安徽": ["合肥", "芜湖", "蚌埠", "马鞍山"],
            "广西": ["南宁", "柳州", "桂林"],
            "河北": ["石家庄", "唐山", "保定", "邯郸"],
            "山西": ["太原", "大同", "长治"],
            "甘肃": ["兰州", "天水"],
            "内蒙古": ["呼和浩特", "包头", "鄂尔多斯"],
            "贵州": ["贵阳", "遵义"],
            "江西": ["南昌", "九江"],
            "海南": ["海口", "三亚"],
            "吉林": ["长春", "吉林"],
            "新疆": ["乌鲁木齐"],
            "西藏": ["拉萨"],
            "青海": ["西宁"],
            "宁夏": ["银川"],
        }
        for prov, cities in province_city_map.items():
            if found_city in cities:
                found_prov = prov
                break

    return found_prov, found_city


def parse_date(date_str: str) -> str:
    """统一日期格式为 YYYY-MM-DD。"""
    if not date_str:
        return ""
    # 已格式化
    m = re.search(r'(\d{4}[-/.]\d{1,2}[-/.]\d{1,2})', date_str)
    if m:
        d = m.group(1).replace("/", "-").replace(".", "-")
        parts = d.split("-")
        return f"{parts[0]}-{int(parts[1]):02d}-{int(parts[2]):02d}"
    # 纯数字格式
    m = re.search(r'(\d{8})', date_str)
    if m:
        d = m.group(1)
        return f"{d[:4]}-{d[4:6]}-{d[6:8]}"
    return date_str[:10]


# ============================================================
# 5. 去重
# ============================================================

def deduplicate(items: List[Dict]) -> List[Dict]:
    """按标题（取前20字）+ 来源去重。"""
    seen = set()
    result = []
    for item in items:
        key = f"{item['title'][:20]}|{item['source']}"
        if key not in seen:
            seen.add(key)
            result.append(item)
    return result


# ============================================================
# 6. Excel 输出
# ============================================================

def make_excel(items: List[Dict], today_str: str) -> str:
    """生成格式化的 Excel 报表。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "公告列表"

    # ── 表头 ──
    headers = [
        "序号", "公告类型", "来源", "省份", "地市",
        "项目名称", "种类", "预算(万)", "公告日期",
        "报名/反馈截止", "投标日期", "报名费(元)", "保证金(万)", "网址",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # ── 数据行 ──
    data_font = Font(size=10)
    link_font = Font(size=10, color="0563C1", underline="single")
    data_alignment = Alignment(vertical="center", wrap_text=False)

    for row_idx, item in enumerate(items, 2):
        row_data = [
            row_idx - 1,                          # 序号
            item.get("notice_type", "招标公告"),   # 公告类型
            item["source"],                        # 来源
            item["province"],                      # 省份
            item["city"],                          # 地市
            item["title"],                         # 项目名称
            item["category"],                      # 种类（赛道）
            item.get("budget", ""),                # 预算(万)
            item.get("announce_date", ""),         # 公告日期
            "",                                    # 报名截止（从详情提取，暂留空）
            "",                                    # 投标日期（从详情提取，暂留空）
            "",                                    # 报名费
            "",                                    # 保证金
            item["detail_url"],                    # 网址
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = link_font if col_idx == 14 else data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    # ── 列宽自适应 ──
    col_widths = {
        1: 6,    # 序号
        2: 14,   # 公告类型
        3: 8,    # 来源
        4: 8,    # 省份
        5: 10,   # 地市
        6: 60,   # 项目名称
        7: 16,   # 种类
        8: 10,   # 预算
        9: 12,   # 公告日期
        10: 14,  # 报名截止
        11: 12,  # 投标日期
        12: 10,  # 报名费
        13: 10,  # 保证金
        14: 50,  # 网址
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 冻结首行 + 自动筛选 ──
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{len(items) + 1}"

    # ── 保存 ──
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"三大运营商广告类招标公告_{today_str}.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(str(filepath))
    logger.info(f"Excel 已保存: {filepath}")

    # 复制到桌面
    desktop_path = DESKTOP_DIR / filename
    try:
        wb.save(str(desktop_path))
        logger.info(f"桌面副本已保存: {desktop_path}")
    except Exception as e:
        logger.warning(f"桌面保存失败: {e}")

    return str(filepath)


# ============================================================
# 7. 主流程
# ============================================================

def main():
    today_str = date.today().strftime("%Y-%m-%d")
    today_for_file = date.today().strftime("%Y-%m-%d")
    logger.info(f"╔{'═' * 60}╗")
    logger.info(f"║  三大运营商广告类招标公告 — 每日爬取")
    logger.info(f"║  日期: {today_str}")
    logger.info(f"╚{'═' * 60}╝")

    # ── 第1步：三大平台爬取 ──
    all_candidates = []
    try:
        items = scrape_b2b_10086()
        all_candidates.extend(items)
        logger.info(f"  ✅ 移动: {len(items)} 条")
    except Exception as e:
        logger.error(f"  ❌ 移动爬取失败: {e}")

    try:
        items = scrape_telecom()
        all_candidates.extend(items)
        logger.info(f"  ✅ 电信: {len(items)} 条")
    except Exception as e:
        logger.error(f"  ❌ 电信爬取失败: {e}")

    try:
        items = scrape_unicom()
        all_candidates.extend(items)
        logger.info(f"  ✅ 联通: {len(items)} 条")
    except Exception as e:
        logger.error(f"  ❌ 联通爬取失败: {e}")

    logger.info(f"\n📦 三大平台原始候选总计: {len(all_candidates)} 条")

    if not all_candidates:
        logger.warning("⚠️ 无任何候选数据，输出空 Excel")
        make_excel([], today_for_file)
        return

    # ── 第2步：关键词粗筛 ──
    logger.info("\n🔍 第1级过滤：关键词粗筛...")
    coarse_filtered = []
    for item in all_candidates:
        result = keyword_filter(item["title"])
        if result["is_ad_candidate"]:
            item["_keyword_category"] = result["category"]
            item["_keyword_reason"] = result["reason"]
            coarse_filtered.append(item)
        else:
            logger.debug(f"  关键词排除: {item['title'][:50]} → {result['reason']}")

    logger.info(f"  粗筛通过: {len(coarse_filtered)}/{len(all_candidates)} 条")

    if not coarse_filtered:
        logger.warning("⚠️ 粗筛无候选，输出空 Excel")
        make_excel([], today_for_file)
        return

    # ── 第3步：LLM 精判 ──
    # 优化策略：
    #   - 有明确赛道（非"其他营销类"）且 ≥2 个安全词 → 跳过 LLM，直接采纳
    #   - 仅"其他营销类"或 <2 个安全词 → 送 LLM 验证
    logger.info("\n🤖 第2级过滤：LLM 精判...")
    final_items = []
    ad_count = 0
    non_ad_count = 0
    skip_llm_count = 0
    all_llm_items = []

    for idx, item in enumerate(coarse_filtered, 1):
        title = item["title"]
        kw_cat = item.get("_keyword_category", "")
        kw_reason = item.get("_keyword_reason", "")
        safety_hits = len(_match_keywords(title, SAFETY_KEYWORDS))

        # 判断是否需要 LLM
        needs_llm = (kw_cat == "其他营销类" or safety_hits < 2)

        if not needs_llm:
            # 强匹配：直接采纳
            item["category"] = kw_cat
            item["llm_reason"] = "关键词强匹配,跳过LLM"
            final_items.append(item)
            ad_count += 1
            skip_llm_count += 1
            logger.debug(f"  [{idx}/{len(coarse_filtered)}] ⏭️ 跳过LLM: {title[:40]} → {kw_cat}")
            continue

        # 需要 LLM 验证
        all_llm_items.append((idx, item))

    # 批量 LLM 判定
    logger.info(f"  LLM跳过(强匹配): {skip_llm_count}, 待LLM判定: {len(all_llm_items)}/{len(coarse_filtered)}")
    for batch_idx, (orig_idx, item) in enumerate(all_llm_items, 1):
        title_short = item["title"][:40]
        logger.info(f"  LLM[{batch_idx}/{len(all_llm_items)}]: {title_short}...")
        llm_result = llm_classify(item["title"])

        if llm_result["is_ad"]:
            cat = llm_result["category"] or item.get("_keyword_category", "")
            item["category"] = cat
            item["llm_reason"] = llm_result["reason"]
            final_items.append(item)
            ad_count += 1
            logger.info(f"    ✅ 广告类 → {cat} ({llm_result['reason']})")
        else:
            non_ad_count += 1
            logger.info(f"    ❌ 非广告: {llm_result['reason']}")

    logger.info(f"\n  LLM判定: 广告类={ad_count}(跳过{skip_llm_count}), 非广告={non_ad_count}")

    # ── 第4步：去重 ──
    logger.info("\n🔗 去重...")
    final_items = deduplicate(final_items)
    logger.info(f"  去重后: {len(final_items)} 条")

    # ── 第5步：补充字段（省份/城市/日期格式化）──
    logger.info("\n📋 补充字段...")
    for item in final_items:
        prov, city = extract_province(item["title"], item.get("province", ""))
        item["province"] = prov
        item["city"] = city
        item["announce_date"] = parse_date(item.get("publish_date", today_str))
        if not item.get("category"):
            item["category"] = item.get("_keyword_category", "其他营销类")

    # ── 第6步：排序（按来源+日期）──
    final_items.sort(key=lambda x: (x["source"], x.get("publish_date", "")), reverse=True)

    # ── 第7步：输出 Excel ──
    logger.info(f"\n📊 生成 Excel...")
    filepath = make_excel(final_items, today_for_file)

    # ── 结果摘要 ──
    logger.info(f"\n{'─' * 60}")
    logger.info(f"✅ 完成！共 {len(final_items)} 条广告类招标公告")
    from collections import Counter
    cat_counts = Counter(item.get("category", "未分类") for item in final_items)
    for cat, count in cat_counts.most_common():
        logger.info(f"  • {cat}: {count} 条")
    logger.info(f"\n📁 输出文件: {filepath}")
    logger.info(f"{'─' * 60}")

    # 打印汇总供 cron 交付
    print(f"\n{'=' * 60}")
    print(f"三大运营商广告类招标公告 — {today_str}")
    print(f"{'=' * 60}")
    print(f"爬取汇总:")
    print(f"  ✅ 移动: {sum(1 for x in all_candidates if x['source']=='移动')} 条原始")
    print(f"  ✅ 电信: {sum(1 for x in all_candidates if x['source']=='电信')} 条原始")
    print(f"  ✅ 联通: {sum(1 for x in all_candidates if x['source']=='联通')} 条原始")
    print(f"  🔍 粗筛通过: {len(coarse_filtered)} 条")
    print(f"  🤖 LLM确认广告类: {ad_count} 条 (强匹配跳过LLM: {skip_llm_count})")
    print(f"  ❌ LLM排除: {non_ad_count} 条")
    print(f"  🔗 去重后: {len(final_items)} 条")
    print(f"\n赛道分布:")
    for cat, count in Counter(item.get("category", "未分类") for item in final_items).most_common():
        print(f"  {cat}: {count} 条")
    print(f"\n📁 输出: {filepath}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
