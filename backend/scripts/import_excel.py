"""
标中宝 — Excel 数据导入脚本

将「致合项目查询汇总2026.6.24.xlsx」导入 PostgreSQL。
支持：
  - 自动字段映射
  - 广东移动智能过滤
  - 广告关键词分类
  - 断点续传（跳过已存在URL）
  - 两 Sheet 合并导入

用法:
    python scripts/import_excel.py
    python scripts/import_excel.py --file "C:/path/to/file.xlsx"
    python scripts/import_excel.py --dry-run          # 预览不写入
"""

import os
import sys
import re
import json
import hashlib
import argparse
import logging
from datetime import datetime, date, time, timedelta
from typing import Optional, Any

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.services.keyword_filter import filter_advertisement_projects

logger = logging.getLogger(__name__)

# 默认 Excel 路径
DEFAULT_EXCEL = r"C:\Users\18826\Desktop\致合项目查询汇总2026.6.24.xlsx"


# ============================================================
# Excel 日期转换
# ============================================================

def excel_serial_to_date(serial: Any) -> Optional[str]:
    """将 Excel 日期序列号转为 ISO 日期字符串。"""
    if serial is None:
        return None
    if isinstance(serial, datetime):
        return serial.strftime("%Y-%m-%d")
    if isinstance(serial, date):
        return serial.strftime("%Y-%m-%d")
    if isinstance(serial, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(serial))).strftime("%Y-%m-%d")
        except Exception:
            return str(serial)
    return str(serial)


def excel_time_to_str(val: Any) -> Optional[str]:
    """将 Excel 时间值转为 HH:MM 字符串。"""
    if val is None:
        return None
    if isinstance(val, time):
        return val.strftime("%H:%M")
    if isinstance(val, timedelta):
        total_sec = int(val.total_seconds())
        h, m = divmod(total_sec // 60, 60)
        return f"{h:02d}:{m:02d}"
    if isinstance(val, datetime):
        return val.strftime("%H:%M")
    s = str(val).strip()
    if re.match(r'^\d{1,2}:\d{2}', s):
        return s
    return None


# ============================================================
# 广东移动过滤器
# ============================================================

GD_KEYWORDS = [
    "广东移动", "中国移动广东", "广州移动", "深圳移动",
    "东莞移动", "佛山移动", "珠海移动", "中山移动", "惠州移动",
    "汕头移动", "江门移动", "湛江移动", "茂名移动", "肇庆移动",
    "梅州移动", "汕尾移动", "河源移动", "阳江移动", "清远移动",
    "潮州移动", "揭阳移动", "云浮移动", "韶关移动",
    "中国移动通信集团广东", "广东有限公司",
]

GD_PROVINCES = ["广东", "广东省"]
GD_CITIES = ["广州", "深圳", "东莞", "佛山", "珠海", "中山",
              "惠州", "汕头", "江门", "湛江", "茂名", "肇庆",
              "梅州", "汕尾", "河源", "阳江", "清远", "潮州",
              "揭阳", "云浮", "韶关"]


def is_gd_mobile(title: str, province: str, city: str) -> bool:
    """判断是否为广东移动相关项目。"""
    text = f"{title} {province or ''} {city or ''}"
    for kw in GD_KEYWORDS:
        if kw in text:
            return True
    # 省份或地市在广东 + 标题含广告类关键词
    loc = f"{province or ''} {city or ''}"
    is_gd = any(r in loc for r in GD_PROVINCES + GD_CITIES)
    has_ad = any(k in title for k in ["广告", "品牌", "营销", "宣传", "活动",
                                        "设计", "制作", "视频", "新媒体", "投放",
                                        "策划", "创意", "展会", "论坛", "发布会",
                                        "党群", "党建", "工会", "培训", "物料"])
    return is_gd and has_ad


# ============================================================
# Excel → ORM 映射
# ============================================================

def row_to_dict(row_data: dict, sheet_name: str = "") -> dict:
    """
    将 Excel 行映射为 Announcement 字段。

    Excel 列 → Model 字段:
      日期 → announce_date
      行业 → industry (招标单位)
      省份 → province
      地市 → city
      项目名称 → title
      种类 → project_category
      预算金额（万元）→ budget
      网址 → source_url
      报名截止日期 → deadline
      截止时间 → deadline_time
      投标日期 → bid_date
      时间 → bid_time
      报名费 → registration_fee
      保证金 → deposit
      备注 → remark
    """
    title = str(row_data.get("项目名称") or "").strip()
    province = str(row_data.get("省份") or "").strip()
    city = str(row_data.get("地市") or "").strip()
    industry = str(row_data.get("行业") or "").strip()

    # 如果行业为空，根据标题推断招标单位
    if not industry:
        industry = _guess_bidder(title)

    budget_raw = row_data.get("预算金额\n（万元）")
    budget = None
    if budget_raw is not None:
        try:
            budget = float(budget_raw)
        except (ValueError, TypeError):
            pass

    registration_fee = None
    fee_raw = row_data.get("报名费")
    if fee_raw is not None:
        try:
            registration_fee = float(fee_raw)
        except (ValueError, TypeError):
            pass

    deposit = None
    dep_raw = row_data.get("保证金")
    if dep_raw is not None:
        try:
            deposit = float(dep_raw)
        except (ValueError, TypeError):
            pass

    deadline = row_data.get("报名截止日期")
    deadline_str = excel_serial_to_date(deadline)
    deadline_time_str = excel_time_to_str(row_data.get("截止时间"))

    bid_date = row_data.get("投标日期")
    bid_date_str = excel_serial_to_date(bid_date)
    bid_time_str = excel_time_to_str(row_data.get("时间"))

    # 处理"军队采购网"Sheet 的列名差异
    if sheet_name == "军队采购网":
        alt_deadline = row_data.get("报名日期")
        if alt_deadline:
            deadline_str = excel_serial_to_date(alt_deadline)

    return {
        "title": title,
        "industry": industry,
        "province": province,
        "city": city,
        "project_category": str(row_data.get("种类") or "").strip(),
        "budget": budget,
        "source_url": str(row_data.get("网址") or "").strip(),
        "announce_date": excel_serial_to_date(row_data.get("日期")),
        "deadline": deadline_str,
        "deadline_time": deadline_time_str,
        "bid_date": bid_date_str,
        "bid_time": bid_time_str,
        "registration_fee": registration_fee,
        "deposit": deposit,
        "remark": str(row_data.get("备注") or "").strip(),
    }


def _guess_bidder(title: str) -> str:
    """从标题推断招标单位。"""
    if "广东移动" in title or "中国移动广东" in title:
        for city_name in GD_CITIES:
            if city_name in title:
                return f"中国移动通信集团广东有限公司{city_name}分公司"
        return "中国移动通信集团广东有限公司"
    return ""


# ============================================================
# 导入器
# ============================================================

class ExcelImporter:
    def __init__(self, excel_path: str, dry_run: bool = False):
        self.excel_path = excel_path
        self.dry_run = dry_run
        self.stats = {
            "total_rows": 0,
            "gd_mobile": 0,
            "ad_classified": 0,
            "non_ad": 0,
            "imported": 0,
            "skipped_duplicate": 0,
            "errors": 0,
        }
        self.existing_urls = set()

    def _load_existing_urls(self):
        """从数据库加载已存在的 URL 用于去重。"""
        try:
            import asyncio
            from sqlalchemy import select
            from app.db.session import engine, get_db
            from app.models.announcement import Announcement

            async def _load():
                async with engine.begin() as conn:
                    result = await conn.execute(select(Announcement.source_url))
                    for row in result:
                        url = row[0]
                        if url:
                            self.existing_urls.add(hashlib.md5(url.encode()).hexdigest())

            asyncio.run(_load())
            logger.info(f"📂 数据库已有 {len(self.existing_urls)} 条记录")
        except Exception as e:
            logger.warning(f"无法连接数据库加载已有URL: {e}，将全部导入")

    def run(self) -> dict:
        """执行导入。"""
        logger.info("=" * 60)
        logger.info(f"📥 标中宝 Excel 导入")
        logger.info(f"  文件: {self.excel_path}")
        logger.info(f"  模式: {'试运行（不写入）' if self.dry_run else '正式导入'}")
        logger.info("=" * 60)

        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        logger.info(f"📑 工作表: {wb.sheetnames}")

        if not self.dry_run:
            self._load_existing_urls()

        all_rows = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [str(ws.cell(1, c).value or "") for c in range(1, ws.max_column + 1)]
            logger.info(f"\n📋 Sheet '{sheet_name}': {ws.max_row - 1} 行")

            sheet_count = 0
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx in range(1, ws.max_column + 1):
                    col_name = headers[col_idx - 1]
                    row_data[col_name] = ws.cell(row_idx, col_idx).value

                item = row_to_dict(row_data, sheet_name)
                title = item["title"]

                if not title or len(title) < 5:
                    continue

                self.stats["total_rows"] += 1

                # 广东移动过滤
                if not is_gd_mobile(title, item["province"], item["city"]):
                    continue

                self.stats["gd_mobile"] += 1

                # 广告关键词分类
                filter_result = filter_advertisement_projects(title, "")
                item["is_ad"] = filter_result["is_ad"]
                item["matched_keywords"] = filter_result.get("matched_keywords", [])
                if filter_result["is_ad"]:
                    self.stats["ad_classified"] += 1
                    if not item["project_category"]:
                        item["project_category"] = filter_result.get("category", "")
                else:
                    self.stats["non_ad"] += 1

                all_rows.append(item)
                sheet_count += 1

            logger.info(f"  广东移动相关: {sheet_count} 条")

        logger.info(f"\n📊 过滤统计:")
        logger.info(f"  总行数:        {self.stats['total_rows']}")
        logger.info(f"  广东移动相关:  {self.stats['gd_mobile']}")
        logger.info(f"  广告类:        {self.stats['ad_classified']}")
        logger.info(f"  非广告类:      {self.stats['non_ad']}")

        if self.dry_run:
            self._save_preview(all_rows)
        else:
            self._import_to_db(all_rows)

        return self.stats

    def _save_preview(self, rows: list):
        """试运行：保存 JSON 预览。"""
        output_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            "output", "excel_import_preview.json"
        )
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump({
                "stats": self.stats,
                "items": rows[:50],  # 仅前 50 条
                "total": len(rows),
            }, f, ensure_ascii=False, indent=2)
        logger.info(f"💾 预览已保存: {output_path}")

    def _import_to_db(self, rows: list):
        """写入 PostgreSQL 数据库（psycopg2 直连）。"""
        import psycopg2
        import os

        db_host = os.getenv("BZB_DB_HOST", "localhost")
        db_port = os.getenv("BZB_DB_PORT", "5432")
        db_user = os.getenv("BZB_DB_USER", "postgres")
        db_pass = os.getenv("BZB_DB_PASS", "postgres")
        db_name = os.getenv("BZB_DB_NAME", "biaozhongbao")

        conn = psycopg2.connect(
            host=db_host, port=db_port, user=db_user,
            password=db_pass, dbname=db_name
        )
        conn.autocommit = True  # 每条独立事务，单条失败不影响其他
        cur = conn.cursor()

        try:
            # 预加载已有 purchaser
            cur.execute("SELECT name, id FROM purchasers")
            purchaser_cache = {row[0]: row[1] for row in cur.fetchall()}
            logger.info(f"📂 已有 {len(purchaser_cache)} 个采购方")

            # 预加载已有 URL
            cur.execute("SELECT source_url FROM announcements WHERE source_url IS NOT NULL AND source_url != ''")
            existing = {hashlib.md5((r[0] or "").encode()).hexdigest() for r in cur.fetchall()}
            logger.info(f"📂 已有 {len(existing)} 条公告URL")

            count = 0
            for item in rows:
                url = item.get("source_url", "")
                url_hash = hashlib.md5(url.encode()).hexdigest() if url else ""

                if url_hash and url_hash in existing:
                    self.stats["skipped_duplicate"] += 1
                    continue

                try:
                    # 解析日期
                    deadline_dt = None
                    if item.get("deadline"):
                        try:
                            dl = item["deadline"]
                            tm = item.get("deadline_time") or "00:00"
                            deadline_dt = datetime.strptime(f"{dl} {tm}", "%Y-%m-%d %H:%M")
                        except ValueError:
                            try:
                                deadline_dt = datetime.strptime(item["deadline"], "%Y-%m-%d")
                            except ValueError:
                                pass

                    announce_dt = None
                    if item.get("announce_date"):
                        try:
                            announce_dt = datetime.strptime(item["announce_date"], "%Y-%m-%d").date()
                        except ValueError:
                            pass

                    bid_dt = None
                    if item.get("bid_date"):
                        try:
                            bid_dt = datetime.strptime(item["bid_date"], "%Y-%m-%d").date()
                        except ValueError:
                            pass

                    # purchaser
                    purchaser_name = item.get("industry") or "未知采购方"
                    if purchaser_name not in purchaser_cache:
                        level = "省公司" if ("有限公司" in purchaser_name and "分公司" not in purchaser_name) else "地市公司"
                        cur.execute(
                            "INSERT INTO purchasers (name, level, region) VALUES (%s, %s, %s) RETURNING id",
                            (purchaser_name, level, item.get("province") or item.get("city") or "广东")
                        )
                        purchaser_cache[purchaser_name] = cur.fetchone()[0]
                    purchaser_id = purchaser_cache[purchaser_name]

                    # 插入公告
                    cur.execute("""
                        INSERT INTO announcements 
                        (title, purchaser_id, purchaser_level, procurement_method, 
                         budget, project_category, announce_date, deadline, 
                         source_url, industry, province, city,
                         bid_date, registration_fee, deposit, remark)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        item["title"],
                        purchaser_id,
                        "省公司" if ("有限公司" in purchaser_name and "分公司" not in purchaser_name) else "地市公司",
                        "公开招标",
                        item.get("budget"),
                        item.get("project_category") or "",
                        announce_dt,
                        deadline_dt,
                        item.get("source_url") or "",
                        purchaser_name,
                        item.get("province") or "",
                        item.get("city") or "",
                        bid_dt,
                        item.get("registration_fee"),
                        item.get("deposit"),
                        item.get("remark") or "",
                    ))

                    if url_hash:
                        existing.add(url_hash)
                    count += 1
                    self.stats["imported"] += 1

                    if count % 50 == 0:
                        logger.info(f"  已导入 {count} 条...")

                except Exception as e:
                    logger.debug(f"插入失败: {item.get('title','')[:60]}... | {e}")
                    self.stats["errors"] += 1

            logger.info(f"✅ 成功导入 {count} 条")

        finally:
            cur.close()
            conn.close()


# ============================================================
# 入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="标中宝 - Excel 数据导入")
    parser.add_argument("--file", type=str, default=DEFAULT_EXCEL, help="Excel 文件路径")
    parser.add_argument("--dry-run", action="store_true", help="试运行，不写入数据库")
    parser.add_argument("--verbose", "-v", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    importer = ExcelImporter(args.file, dry_run=args.dry_run)
    stats = importer.run()

    print(f"\n{'='*60}")
    print(f"[导入完成]")
    print(f"  总行数:       {stats['total_rows']}")
    print(f"  广东移动:     {stats['gd_mobile']}")
    print(f"  广告类:       {stats['ad_classified']}")
    print(f"  非广告类:     {stats['non_ad']}")
    print(f"  已导入:       {stats['imported']}")
    print(f"  跳过重复:     {stats['skipped_duplicate']}")
    print(f"  错误:         {stats['errors']}")


if __name__ == "__main__":
    main()
